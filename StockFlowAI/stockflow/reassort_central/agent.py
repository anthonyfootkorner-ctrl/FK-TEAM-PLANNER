#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Agent Réassort Multi-Boutiques v1.0
====================================
Cet agent se comporte comme un approvisionneur :
- Il analyse les ventes et les stocks de toutes les boutiques
- Il calcule la vitesse d'écoulement par article/taille/boutique
- Il identifie les risques de rupture avant qu'ils arrivent
- Il propose des transferts depuis CENTRAL (réserve) vers les boutiques
- Il alloue intelligemment le stock réserve par ordre de priorité
- Il évite le sur-stockage en plafonnant à la couverture cible

Usage:
  python agent_reassort_multiboutiques.py
      → sélecteur de fichiers graphique

  python agent_reassort_multiboutiques.py --stock stocks.csv --ventes ventes.csv
      → mode automatique sans interface

  python agent_reassort_multiboutiques.py --stock stocks.csv --ventes ventes.csv
        --reserve CENTRAL --target-days 28 --output reassort.xlsx

Dépendances: pandas openpyxl
Installation: python -m pip install pandas openpyxl
"""

from __future__ import annotations

import argparse
import math
import re
import sys
import traceback
from datetime import datetime
from pathlib import Path

try:
    import pandas as pd
except ImportError:
    print("ERREUR : pandas n'est pas installé.")
    print("Lance : python -m pip install pandas openpyxl")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.formatting.rule import FormulaRule
except ImportError:
    print("ERREUR : openpyxl n'est pas installé.")
    print("Lance : python -m pip install pandas openpyxl")
    sys.exit(1)

# ──────────────────────────────────────────────────────────────────────────────
# Paramètres par défaut
# ──────────────────────────────────────────────────────────────────────────────

DEFAULT_TARGET_DAYS = 14   # Couverture cible en jours
DEFAULT_SALES_DAYS  = 35   # Période de référence ventes
DEFAULT_RESERVE     = "CENTRAL"

# Plancher de stock : toute taille qui se vend (≥ MIN_STOCK_FLOOR_SALES ventes sur
# la période) doit avoir au moins MIN_STOCK_FLOOR pièces en boutique après réassort.
# Évite qu'une taille qui tourne reste à 1 exemplaire (rupture dès la vente suivante).
MIN_STOCK_FLOOR       = 2   # Stock minimum garanti par taille qui se vend
MIN_STOCK_FLOOR_SALES = 1   # Ventes sur 35 j à partir desquelles le plancher s'applique
PROMO_EXEMPTIONS    = {"WEB"}   # Ces boutiques ignorent le filtre promo
SAISON_EXEMPTIONS   = {"WEB"}   # Ces boutiques reçoivent toutes saisons confondues

# Grille cœur (vêtements) — règle : 3 tailles cœur couvertes minimum avant d'envoyer.
# XL/XXL sont des tailles adultes mais ne comptent PAS pour valider la grille cœur.
MENS_SIZES          = {"XS", "S", "M", "L"}                  # grille cœur (jugée pour la complétude)
ADULT_SIZES         = {"XS", "S", "M", "L", "XL", "XXL"}     # toutes tailles vêtement adulte
MIN_COVERED_MENS    = 3    # Nb de tailles cœur (XS/S/M/L) à avoir dispo pour valider le transfert

# Pointures cœur chaussures — règle : 3 pointures dans la plage 41–45 couvertes avant d'envoyer.
# La plage est inclusive et gère les demi-tailles (41.5, 42,5, 44,5…), qui comptent
# comme des pointures cœur distinctes.
# Les pointures hors cœur (≤ 40.5, ≥ 45.5) sont des « extra » : envoyées seulement
# si le cœur 41-45 est complet, jamais seules (même logique que XL/XXL en vêtement).
SHOES_CORE_RANGE    = (41.0, 45.0)   # (min, max) inclusif
SHOES_ALL_RANGE     = (30.0, 50.0)   # plage large pour reconnaître une chaussure
MIN_COVERED_SHOES   = 3              # Nb de pointures distinctes dans la plage cœur à avoir dispo

# ──────────────────────────────────────────────────────────────────────────────
# Utilitaires
# ──────────────────────────────────────────────────────────────────────────────

def _size_to_float(taille: str) -> float | None:
    """
    Convertit une taille numérique en float, quelle que soit la notation :
    '42' → 42.0 | '42.5' → 42.5 | '42,5' → 42.5
    Retourne None si la taille n'est pas numérique (ex : 'M', 'TU', 'S/M').
    """
    try:
        return float(str(taille).replace(",", ".").strip())
    except (ValueError, AttributeError):
        return None


def _in_shoe_range(taille: str, lo: float = SHOES_CORE_RANGE[0], hi: float = SHOES_CORE_RANGE[1]) -> bool:
    """Retourne True si la taille est une pointure dans la plage [lo, hi] (inclus)."""
    v = _size_to_float(taille)
    return v is not None and lo <= v <= hi


def _is_shoe_size(taille: str) -> bool:
    """True si la taille est une pointure de chaussure (numérique dans la plage large),
    y compris hors cœur (40, 41, 45, 46…). Sert à reconnaître un article chaussure."""
    return _in_shoe_range(taille, SHOES_ALL_RANGE[0], SHOES_ALL_RANGE[1])


def _is_core_size(taille: str) -> bool:
    """True si la taille est une taille cœur : XS/S/M/L (vêtement) ou 41-45 (chaussure).
    Sert au plancher de stock minimum (qui ne s'applique qu'aux tailles cœur)."""
    if (taille or "").strip().upper() in MENS_SIZES:
        return True
    return _in_shoe_range(taille)  # cœur chaussures 41-45


def _saison_year(s: str) -> int | None:
    """Extrait la plus grande année (2 chiffres) d'une chaîne saison.
    Exemples : '26 Q1'→26, 'AH23'→23, 'Q1 26'→26, '25/26'→26, '2022'→22.
    Retourne None si aucune année détectable (PERMANENT, ITM…).
    """
    import re
    nums = [int(m) for m in re.findall(r'\d{2}', str(s)) if 15 <= int(m) <= 99]
    return max(nums) if nums else None


def clean_col(name: object) -> str:
    return str(name).replace("﻿", "").replace('"', "").strip()


def to_number(series: pd.Series) -> pd.Series:
    return pd.to_numeric(
        series.astype(str)
        .str.replace(" ", " ", regex=False)
        .str.replace(" ", "", regex=False)
        .str.replace(",", ".", regex=False)
        .str.replace(r"[^0-9.\-]", "", regex=True),
        errors="coerce",
    ).fillna(0)


def read_file(path: Path) -> pd.DataFrame:
    """Lit un CSV ou Excel, gère les encodages courants."""
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        df = pd.read_excel(path, dtype=str)
    elif suffix == ".xls":
        try:
            df = pd.read_excel(path, dtype=str)
        except Exception:
            try:
                df = pd.read_csv(path, sep="\t", encoding="latin1", dtype=str)
            except Exception:
                df = pd.read_csv(path, sep=None, engine="python", encoding="latin1", dtype=str)
    else:
        for enc in ["utf-8-sig", "utf-8", "latin1"]:
            try:
                df = pd.read_csv(path, dtype=str, encoding=enc)
                break
            except UnicodeDecodeError:
                continue
        else:
            raise ValueError(f"Impossible de lire {path}")
    df.columns = [clean_col(c) for c in df.columns]
    return df


def find_col(df: pd.DataFrame, candidates: list[str]) -> str | None:
    """Trouve la première colonne correspondant à l'une des clés candidates."""
    norm = {clean_col(c).lower(): c for c in df.columns}
    for cand in candidates:
        key = clean_col(cand).lower()
        if key in norm:
            return norm[key]
    return None


# ──────────────────────────────────────────────────────────────────────────────
# Chargement des données
# ──────────────────────────────────────────────────────────────────────────────

def load_stock(path: Path) -> pd.DataFrame:
    """
    Charge le fichier de stock multi-boutiques.
    Colonnes attendues : Code_Origine, BarCode V2, Taille, Total Stock
    """
    df = read_file(path)
    boutique_col = find_col(df, ["Code_Origine", "Magasin", "Boutique", "Code_Magasin"])
    barcode_col  = find_col(df, ["BarCode V2", "Barcode", "Code barre", "Gencod", "GenCod", "EAN"])
    taille_col   = find_col(df, ["Taille", "Size", "Pointure"])
    stock_col    = find_col(df, ["Total Stock", "Stock", "Quantité disponible", "Quantite disponible", "Qte"])

    missing = []
    if not boutique_col: missing.append("boutique (Code_Origine)")
    if not barcode_col:  missing.append("barcode (BarCode V2)")
    if not taille_col:   missing.append("taille")
    if not stock_col:    missing.append("stock (Total Stock)")
    if missing:
        raise ValueError(f"Colonnes introuvables dans le fichier stock : {missing}\nColonnes présentes : {list(df.columns)}")

    out = pd.DataFrame({
        "boutique": df[boutique_col].astype(str).str.strip(),
        "barcode":  df[barcode_col].astype(str).str.strip(),
        "taille":   df[taille_col].astype(str).str.strip(),
        "stock":    to_number(df[stock_col]),
    })
    out = out[out["barcode"].notna() & (out["barcode"] != "") & (out["barcode"] != "nan")]
    return out.groupby(["boutique", "barcode", "taille"], as_index=False)["stock"].sum()


def load_ventes(path: Path, default_sales_days: int) -> tuple[pd.DataFrame, str, str, int]:
    """
    Charge le fichier de ventes multi-boutiques.
    Colonnes attendues : Code_Origine, BarCode V2, Taille, [Marque Gp, Saison,] Jours dans Date, Total QteVenteRetail
    """
    df = read_file(path)
    boutique_col = find_col(df, ["Code_Origine", "Magasin", "Boutique"])
    barcode_col  = find_col(df, ["BarCode V2", "Barcode", "Code barre", "Gencod", "GenCod"])
    taille_col   = find_col(df, ["Taille", "Size", "Pointure"])
    qty_col      = find_col(df, ["Total QteVenteRetail", "QteVenteRetail", "Quantite", "Quantité", "Ventes", "Qte"])
    date_col     = find_col(df, ["Jours dans Date", "Date", "Jour"])
    marque_col   = find_col(df, ["Marque Gp", "Marque", "Brand"])
    saison_col   = find_col(df, ["Saison", "Season"])
    prix_col     = find_col(df, ["PrixVente", "Prix vente", "Prix_vente", "PV"])
    mt_col       = find_col(df, ["MtVenteRetailTTC", "MontantVente", "CA TTC", "MtVente"])

    missing = []
    if not boutique_col: missing.append("boutique (Code_Origine)")
    if not barcode_col:  missing.append("barcode (BarCode V2)")
    if not taille_col:   missing.append("taille")
    if not qty_col:      missing.append("quantité (Total QteVenteRetail)")
    if missing:
        raise ValueError(f"Colonnes introuvables dans le fichier ventes : {missing}\nColonnes présentes : {list(df.columns)}")

    out = pd.DataFrame({
        "boutique": df[boutique_col].astype(str).str.strip(),
        "barcode":  df[barcode_col].astype(str).str.strip(),
        "taille":   df[taille_col].astype(str).str.strip(),
        "qty":      to_number(df[qty_col]),
    })
    if marque_col:
        out["marque"] = df[marque_col].astype(str).str.strip()
    if saison_col:
        out["saison"] = df[saison_col].astype(str).str.strip()
    if prix_col:
        out["prix_vente"] = to_number(df[prix_col])
    if mt_col:
        out["mt_realise"] = to_number(df[mt_col])

    # Période réelle des ventes
    sales_days = default_sales_days
    start_date, end_date = "", ""
    if date_col:
        dates = pd.to_datetime(df[date_col], dayfirst=True, errors="coerce")
        if dates.notna().any():
            start_date = dates.min().strftime("%d/%m/%Y")
            end_date   = dates.max().strftime("%d/%m/%Y")
            actual_days = (dates.max() - dates.min()).days + 1
            if actual_days >= 7:
                sales_days = actual_days

    out = out[out["barcode"].notna() & (out["barcode"] != "") & (out["barcode"] != "nan")]
    return out, start_date, end_date, sales_days


# ──────────────────────────────────────────────────────────────────────────────
# Logique de l'agent approvisionneur
# ──────────────────────────────────────────────────────────────────────────────

def classify_priority(coverage_days: float, qty_35j: float, stock: float) -> tuple[str, str]:
    """Classe un besoin par niveau d'urgence."""
    if stock <= 0:
        return "P1 - Rupture",       "Rupture en boutique — urgent"
    if coverage_days < 7:
        return "P1 - <7 jours",      "Risque rupture imminent"
    if coverage_days < 14:
        return "P2 - <14 jours",     "À réapprovisionner vite"
    if qty_35j >= 5:
        return "P3 - Best-seller",   "Bon vendeur à sécuriser"
    return   "P4 - Opportunité",     "Complément possible"


PRIORITY_ORDER = {
    "P1 - Rupture":     1,
    "P1 - <7 jours":   2,
    "P2 - <14 jours":  3,
    "P3 - Best-seller": 4,
    "P4 - Opportunité": 5,
}



def compute_promo_rates(ventes_df: pd.DataFrame, threshold: float = 0.20) -> pd.DataFrame:
    """
    Calcule le taux de remise moyen pondéré par (barcode, taille).

    Logique :
    - Exclut les lignes avec MtVenteRetailTTC = 0 (retours, cadeaux, erreurs)
    - Prix unitaire réalisé = MtVenteRetailTTC / qty
    - Remise = max(0, (PrixVente - unit_realise) / PrixVente)
      → plafonné à 0 si article vendu au-dessus du catalogue
    - Moyenne pondérée par la quantité vendue

    Retourne un DataFrame : barcode, taille, promo_pct, promo_alerte
    """
    if "prix_vente" not in ventes_df.columns or "mt_realise" not in ventes_df.columns:
        return pd.DataFrame(columns=["barcode", "taille", "promo_pct", "promo_alerte"])

    valid = ventes_df[
        (ventes_df["prix_vente"] > 0) &
        (ventes_df["mt_realise"] > 0) &
        (ventes_df["qty"] > 0)
    ].copy()

    if valid.empty:
        return pd.DataFrame(columns=["barcode", "taille", "promo_pct", "promo_alerte"])

    valid["unit_realise"]  = valid["mt_realise"] / valid["qty"]
    valid["remise_unit"]   = ((valid["prix_vente"] - valid["unit_realise"]) / valid["prix_vente"]).clip(lower=0)
    valid["remise_pond"]   = valid["remise_unit"] * valid["qty"]

    agg = valid.groupby(["barcode", "taille"]).agg(
        remise_sum  = ("remise_pond", "sum"),
        qty_sum     = ("qty",         "sum"),
        prix_moyen  = ("prix_vente",  "mean"),
    ).reset_index()

    agg["promo_pct"]    = (agg["remise_sum"] / agg["qty_sum"]).round(4)
    agg["promo_alerte"] = agg["promo_pct"] > threshold

    return agg[["barcode", "taille", "promo_pct", "promo_alerte", "prix_moyen"]]


def _filter_tailles_isolees(
    proposed:           pd.DataFrame,
    boutiques_stock:    pd.DataFrame,
    ventes_df:          pd.DataFrame,
    reserve:            str = DEFAULT_RESERVE,
    min_taille_share:   float = 0.15,
    taille_exemptions:  set[str] | None = None,
    mens_sizes:         set[str] | None = None,
    min_covered_mens:   int = 3,
    min_covered_shoes:  int = 3,
    min_qty_core:       int = 2,
    reserve_pool:       dict | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Filtre et complète les courbes de tailles rompues.

    Pour chaque (boutique, barcode) dont la courbe est incomplète :
      1. Tente de compléter en ajoutant les tailles manquantes disponibles en réserve
         (même sans besoin calculé — 0 ventes boutique).
      2. Si la courbe est complète après complétion → on envoie tout.
      3. Si la réserve n'a pas assez pour compléter → on exclut (taille isolée).

    Trois règles selon le type d'article :
      A — Homme (XS/S/M/L) : ≥ min_covered_mens tailles couvertes.
      B — Chaussures (41–45 incl. demi-tailles) : ≥ min_covered_shoes pointures couvertes.
      C — Autres : au moins une taille cœur (≥ min_taille_share des ventes) couverte.

    "Couverte" = stock boutique > 0 OU dans les transferts proposés (existants + complétion).
    Les boutiques dans taille_exemptions (défaut : WEB) sont exemptées.

    Retourne : (proposed_clean_avec_completions, taille_seule_exclus)
    """
    exempt = set(taille_exemptions) if taille_exemptions is not None else {"WEB"}
    mens   = set(mens_sizes)        if mens_sizes        is not None else MENS_SIZES

    if proposed.empty:
        return proposed, pd.DataFrame(columns=list(proposed.columns) + ["raison_exclusion"])

    # ── Tailles de chaque barcode présentes dans les ventes réseau ───────────
    barcode_tailles: dict[str, set] = (
        ventes_df.groupby("barcode")["taille"].apply(set).to_dict()
    )

    # ── Tailles cœur par barcode (Règle C) ──────────────────────────────────
    size_sales = (
        ventes_df.groupby(["barcode", "taille"])["qty"].sum()
        .reset_index().rename(columns={"qty": "qty_taille"})
    )
    barcode_total = (
        ventes_df.groupby("barcode")["qty"].sum()
        .reset_index().rename(columns={"qty": "qty_barcode"})
    )
    size_sales = size_sales.merge(barcode_total, on="barcode", how="left")
    size_sales["share"] = size_sales["qty_taille"] / size_sales["qty_barcode"].replace(0, 1)
    core_sizes: dict[str, set] = (
        size_sales[size_sales["share"] >= min_taille_share]
        .groupby("barcode")["taille"].apply(set).to_dict()
    )

    # ── Stock boutique par (boutique, barcode, taille) ───────────────────────
    stock_lookup = (
        boutiques_stock[boutiques_stock["stock"] > 0]
        .set_index(["boutique", "barcode", "taille"])["stock"].to_dict()
    )

    # ── Stock réserve par (barcode, taille) ─────────────────────────────────
    reserve_lookup = (
        boutiques_stock[
            (boutiques_stock["boutique"] == reserve) & (boutiques_stock["stock"] > 0)
        ].set_index(["barcode", "taille"])["stock"].to_dict()
    )

    # ── Méta barcode (marque, saison) pour les lignes de complétion ─────────
    meta_cols = [c for c in ["marque", "saison", "qty_global",
                              "urgence_boutique_j", "_boutique_urgency", "_boutique_volume",
                              "promo_pct"] if c in proposed.columns]
    barcode_meta: dict[str, dict] = {}
    for bc, grp_bc in proposed.groupby("barcode"):
        barcode_meta[bc] = {c: grp_bc[c].iloc[0] for c in meta_cols}

    mask_exclude  = pd.Series(False, index=proposed.index)
    completion_rows: list[dict] = []   # Nouvelles lignes de complétion

    for (boutique, barcode), grp in proposed.groupby(["boutique", "barcode"]):
        if boutique in exempt:
            continue

        proposed_tailles = set(grp["taille"].values)

        def covered(t: str, extra: set | None = None) -> bool:
            """Taille couverte = proposée (originale + complétion) OU en stock boutique."""
            tailles = proposed_tailles | (extra or set())
            return (t in tailles) or (stock_lookup.get((boutique, barcode, t), 0) > 0)

        def reserve_qty(t: str) -> int:
            """Stock disponible en réserve pour cette taille (pool décrémenté si fourni)."""
            key = (barcode, t)
            if reserve_pool is not None:
                avail = int(reserve_pool.get(key, 0))
                return avail
            return int(reserve_lookup.get(key, 0))

        def make_completion(taille: str, qte: int) -> dict:
            """Crée une ligne de transfert de complétion de courbe."""
            # Décrémenter le pool pour respecter le stock réel
            if reserve_pool is not None:
                key = (barcode, taille)
                reserve_pool[key] = max(0, reserve_pool.get(key, 0) - qte)
            row: dict = {
                "boutique":          boutique,
                "barcode":           barcode,
                "taille":            taille,
                "stock":             0,
                "stock_reserve":     reserve_qty(taille),
                "qty_35j":           0,
                "ventes_jour":       0.0,
                "couverture_jours":  0.0,
                "besoin_theorique":  qte,
                "qte_proposee":      qte,
                "priorite":          "P4 - Opportunité",
                "commentaire":       "Complétion courbe tailles",
            }
            row.update(barcode_meta.get(barcode, {}))
            return row

        barcode_known_tailles = barcode_tailles.get(barcode, set())

        # ── Règle A : vêtement adulte (≥2 tailles XS→XXL, dont au moins 1 cœur) ────
        # La grille se juge UNIQUEMENT sur le cœur XS/S/M/L : il faut
        # min_covered_mens tailles cœur couvertes après réassort, sinon on exclut
        # TOUT (y compris XL/XXL) → « XL tout seul = pas de réassort ».
        # XL/XXL ne comptent pas pour la grille et ne sont pas complétés d'office :
        # ils sont simplement conservés (car issus de la demande) si la grille cœur
        # est complète → « si S/M/L déjà là, on envoie aussi le XL ».
        adult_in_barcode = barcode_known_tailles & ADULT_SIZES
        core_in_barcode  = barcode_known_tailles & mens
        if len(adult_in_barcode) >= 2 and len(core_in_barcode) >= 1:
            missing_core = [t for t in mens if not covered(t)]
            extra   = set()
            new_compl = []
            for t in missing_core:
                q = reserve_qty(t)
                if q > 0:
                    qte = min(max(min_qty_core, 1), q)
                    new_compl.append(make_completion(t, qte))
                    extra.add(t)
            n_core = sum(1 for t in mens if covered(t, extra))
            if n_core >= min_covered_mens:
                completion_rows.extend(new_compl)  # grille cœur OK → on garde tout (dont XL/XXL)
            else:
                mask_exclude.loc[grp.index] = True  # grille cœur incomplète → tout exclu
            continue

        # ── Règle B : chaussures (pointure numérique, ≥2 pointures dont ≥1 cœur 41–45) ─
        # La grille se juge sur le cœur 41-45 : il faut min_covered_shoes pointures
        # cœur couvertes après réassort, sinon on exclut TOUT (dont 45/46).
        # Les pointures hors cœur ne sont pas complétées d'office, juste conservées.
        shoe_sizes_barcode = {t for t in barcode_known_tailles if _is_shoe_size(t)}
        core_shoes_barcode = {t for t in barcode_known_tailles if _in_shoe_range(t)}
        if len(shoe_sizes_barcode) >= 2 and len(core_shoes_barcode) >= 1:
            # Valeurs float déjà couvertes dans le cœur 41-45
            covered_vals: set[float] = set()
            for t in core_shoes_barcode:
                if covered(t):
                    v = _size_to_float(t)
                    if v is not None:
                        covered_vals.add(v)

            if len(covered_vals) < min_covered_shoes:
                # Complète UNIQUEMENT dans le cœur 41-45, entières d'abord
                candidates = sorted(
                    core_shoes_barcode,
                    key=lambda t: (_size_to_float(t) % 1 != 0, _size_to_float(t))
                )
                new_compl = []
                extra_vals: set[float] = set(covered_vals)
                for t in candidates:
                    v = _size_to_float(t)
                    if v in extra_vals:
                        continue  # déjà couverte (autre notation)
                    if covered(t):
                        extra_vals.add(v)
                        continue
                    q = reserve_qty(t)
                    if q > 0:
                        qte = min(max(min_qty_core, 1), q)
                        new_compl.append(make_completion(t, qte))
                        extra_vals.add(v)
                    if len(extra_vals) >= min_covered_shoes:
                        break
                if len(extra_vals) >= min_covered_shoes:
                    completion_rows.extend(new_compl)   # complété ✅
                else:
                    mask_exclude.loc[grp.index] = True  # insuffisant → exclure
            continue

        # ── Règle C : autres articles ────────────────────────────────────────
        cores = core_sizes.get(barcode, set())
        if not cores:
            continue

        if not any(covered(t) for t in cores):
            # Tente de compléter avec la première taille cœur disponible en réserve
            new_compl = []
            extra = set()
            for t in sorted(cores):
                q = reserve_qty(t)
                if q > 0:
                    qte = min(max(min_qty_core, 1), q)
                    new_compl.append(make_completion(t, qte))
                    extra.add(t)
                    break  # une taille cœur suffit pour la Règle C
            if extra and any(covered(t, extra) for t in cores):
                completion_rows.extend(new_compl)   # complété ✅
            else:
                mask_exclude.loc[grp.index] = True  # pas de taille cœur dispo → exclure

    proposed_clean      = proposed[~mask_exclude].copy()
    taille_seule_exclus = proposed[mask_exclude].copy()
    if not taille_seule_exclus.empty:
        taille_seule_exclus["raison_exclusion"] = "Courbe tailles rompue"

    # ── Ajout des lignes de complétion ───────────────────────────────────────
    if completion_rows:
        df_compl = pd.DataFrame(completion_rows)
        # Assure que toutes les colonnes de proposed_clean sont présentes
        for col in proposed_clean.columns:
            if col not in df_compl.columns:
                df_compl[col] = None
        df_compl = df_compl[proposed_clean.columns]
        proposed_clean = pd.concat([proposed_clean, df_compl], ignore_index=True)

    return proposed_clean, taille_seule_exclus


def run_agent(
    stock_df:           pd.DataFrame,
    ventes_df:          pd.DataFrame,
    reserve:            str,
    target_days:        int,
    sales_days:         int,
    promo_df:           pd.DataFrame | None = None,
    promo_seuil:        float = 0.20,
    promo_exemptions:   set[str] | None = None,
    min_taille_share:   float = 0.15,
    taille_exemptions:  set[str] | None = None,
    min_qty_core:       int   = 2,
    min_covered_mens:   int   = 3,
    min_covered_shoes:  int   = 3,
    min_saison_year:    int   = 26,
    saison_exemptions:  set[str] | None = None,
    min_stock_floor:    int   = MIN_STOCK_FLOOR,
    floor_min_sales:    int   = MIN_STOCK_FLOOR_SALES,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Cœur de l'agent approvisionneur.

    Étapes :
    1.  Agrège les ventes par (boutique, barcode, taille)
    2.  Calcule la vitesse journalière
    3.  Fusionne avec le stock boutique
    4.  Calcule la couverture et le besoin théorique
    5.  Filtre les articles en promo > seuil (exclus du réassort)
    6.  Vérifie le stock disponible en réserve (CENTRAL)
    7.  Alloue le stock réserve par ordre de priorité (P1 → P4)
        → évite de sur-stocker : plafonne au besoin réel
        → gère la rareté : si CENTRAL est limité, les urgences passent en premier
    8.  Identifie les risques sans réserve
    9.  Filtre les courbes de tailles rompues

    Retourne : (propositions, risques, top_ventes_global, promo_exclus, taille_seule_exclus)
    """

    # ── 0. Filtre saison : exclut les articles antérieurs à min_saison_year ────
    # Les boutiques dans saison_exemptions (WEB) reçoivent toutes saisons.
    _saison_exempt = set(saison_exemptions) if saison_exemptions is not None else SAISON_EXEMPTIONS
    SAISON_EXCLUDES = {"ANCIEN STO", "ANCIEN"}
    if "saison" in ventes_df.columns and min_saison_year > 0:
        def _keep_saison(s: str) -> bool:
            if str(s).strip().upper() in SAISON_EXCLUDES:
                return False
            y = _saison_year(s)
            return y is None or y >= min_saison_year
        # Applique le filtre uniquement sur les boutiques NON exemptées
        mask_saison = (
            ventes_df["boutique"].isin(_saison_exempt) |
            ventes_df["saison"].apply(_keep_saison)
        )
        if (~mask_saison).sum() > 0:
            ventes_df = ventes_df[mask_saison].copy()

    # ── 1. Ventes agrégées par (boutique, barcode, taille) ────────────────────
    group_cols = ["boutique", "barcode", "taille"]
    sales_agg = ventes_df.groupby(group_cols)["qty"].sum().reset_index()
    sales_agg.columns = group_cols + ["qty_35j"]

    # Enrichissement marque / saison (premier enregistrement)
    for extra in ["marque", "saison"]:
        if extra in ventes_df.columns:
            extra_df = ventes_df.groupby(group_cols)[extra].first().reset_index()
            sales_agg = sales_agg.merge(extra_df, on=group_cols, how="left")

    # Ventes globales toutes boutiques (pour identifier bestsellers réseau)
    sales_global = ventes_df.groupby(["barcode", "taille"])["qty"].sum().reset_index()
    sales_global.columns = ["barcode", "taille", "qty_global"]

    # ── 2. Stock boutiques (hors réserve) ─────────────────────────────────────
    boutiques_stock = stock_df[stock_df["boutique"] != reserve].copy()

    # ── 3. Stock réserve ──────────────────────────────────────────────────────
    reserve_stock = (
        stock_df[stock_df["boutique"] == reserve]
        .groupby(["barcode", "taille"], as_index=False)["stock"].sum()
        .rename(columns={"stock": "stock_reserve"})
    )

    # ── 4. Fusion stock boutique + ventes ─────────────────────────────────────
    # outer join pour attraper : articles en stock non vendus ET articles vendus sans stock restant
    merged = boutiques_stock.merge(sales_agg, on=group_cols, how="outer")
    merged["stock"]   = merged["stock"].fillna(0)
    merged["qty_35j"] = merged["qty_35j"].fillna(0)

    # Rattacher le stock réserve
    merged = merged.merge(reserve_stock, on=["barcode", "taille"], how="left")
    merged["stock_reserve"] = merged["stock_reserve"].fillna(0)

    # Rattacher ventes globales réseau
    merged = merged.merge(sales_global, on=["barcode", "taille"], how="left")
    merged["qty_global"] = merged["qty_global"].fillna(0)

    # ── 5. Calculs de couverture et besoin ────────────────────────────────────
    merged["ventes_jour"] = merged["qty_35j"] / sales_days

    merged["couverture_jours"] = merged.apply(
        lambda r: 9999.0 if r["ventes_jour"] <= 0 else round(r["stock"] / r["ventes_jour"], 1),
        axis=1,
    )
    merged["_is_core"] = merged["taille"].apply(_is_core_size)

    def _besoin(r):
        # Besoin par couverture cible
        besoin = math.ceil(target_days * r["ventes_jour"] - r["stock"])
        # Plancher : une taille CŒUR qui se vend doit avoir au moins min_stock_floor pièces
        if r["qty_35j"] >= floor_min_sales and r["_is_core"]:
            besoin = max(besoin, min_stock_floor - r["stock"])
        return max(0, int(besoin))

    merged["besoin_theorique"] = merged.apply(_besoin, axis=1)

    # ── 6. Classification priorité ────────────────────────────────────────────
    prio = merged.apply(
        lambda r: classify_priority(r["couverture_jours"], r["qty_35j"], r["stock"]),
        axis=1,
    )
    merged["priorite"]    = [p[0] for p in prio]
    merged["commentaire"] = [p[1] for p in prio]
    merged["_prio_num"]   = merged["priorite"].map(PRIORITY_ORDER).fillna(9)

    # ── 7. Filtrage : seulement les articles qui ont un besoin ────────────────
    needs = merged[merged["besoin_theorique"] > 0].copy()

    # ── 7b. Filtre promo : exclure les articles trop remisés ──────────────────
    # Si le fichier de ventes contient PrixVente et MtVenteRetailTTC,
    # on exclut du réassort tout article dont la remise moyenne dépasse le seuil.
    if promo_df is not None and not promo_df.empty:
        needs = needs.merge(
            promo_df[["barcode", "taille", "promo_pct", "promo_alerte"]],
            on=["barcode", "taille"], how="left",
        )
    else:
        needs["promo_pct"]    = 0.0
        needs["promo_alerte"] = False
    needs["promo_pct"]    = needs["promo_pct"].fillna(0.0)
    needs["promo_alerte"] = needs["promo_alerte"].fillna(False).infer_objects(copy=False)

    # Boutiques exemptées du filtre promo (WEB, etc.)
    exempt = set(promo_exemptions) if promo_exemptions else PROMO_EXEMPTIONS
    est_exempt = needs["boutique"].isin(exempt)

    # Articles exclus pour promo — uniquement les boutiques non exemptées
    promo_exclus = needs[needs["promo_alerte"] & ~est_exempt].copy()

    # On retire les articles en promo forte SAUF pour les boutiques exemptées
    needs = needs[~needs["promo_alerte"] | est_exempt].copy()

    # ── 7d. Ordre d'allocation : WEB en priorité absolue, puis boutiques classées
    #        par urgence de leur top 30 produits (couverture min sur les 30 meilleures ventes)
    #
    #  Règle 1 : WEB passe toujours en premier — ses besoins consomment CENTRAL avant tout le monde.
    #  Règle 2 : Parmi les autres boutiques, celle dont le produit du top 30 va tomber
    #            en rupture le plus tôt est servie en premier.
    #  Règle 3 : Au sein d'une boutique, les articles les plus urgents (P1 > P2 > …)
    #            passent avant les moins urgents.
    #  Règle 4 : À priorité égale, les articles qui se vendent le plus vite passent en premier.

    # ── Urgence boutique : couverture min top30 + volume comme départage ────────
    # Quand plusieurs boutiques ont déjà tout leur top30 en rupture (couverture=0),
    # celle qui vend le plus sur son top30 passe en premier.
    boutique_urgency: dict[str, float] = {}
    boutique_volume:  dict[str, float] = {}

    for boutique, grp in needs.groupby("boutique"):
        top30 = grp.nlargest(30, "qty_35j")
        if boutique == "WEB":
            boutique_urgency["WEB"] = -1.0   # Valeur sentinelle : WEB avant tout
            boutique_volume["WEB"]  = float(top30["qty_35j"].sum())
        else:
            boutique_urgency[boutique] = float(top30["couverture_jours"].min()) if len(top30) else 9999.0
            boutique_volume[boutique]  = float(top30["qty_35j"].sum())

    needs["_is_web"]           = (needs["boutique"] == "WEB").astype(int)
    needs["_boutique_urgency"] = needs["boutique"].map(boutique_urgency).fillna(9999.0)
    needs["_boutique_volume"]  = needs["boutique"].map(boutique_volume).fillna(0.0)

    # Colonne lisible dans l'export
    needs["urgence_boutique_j"] = needs["_boutique_urgency"].apply(
        lambda x: 0.0 if x <= 0 else round(x, 1)
    )

    # Ordre d'allocation :
    #   1. WEB en tête absolue
    #   2. Boutique dont le top30 a la couverture la plus faible (plus urgente)
    #   3. A urgence égale : boutique avec le plus grand volume sur son top30
    #   4. Au sein d'une boutique : articles les plus urgents (P1 > P2 ...)
    #   5. A priorité égale : articles les plus rapides à vendre
    needs = needs.sort_values(
        ["_is_web", "_boutique_urgency", "_boutique_volume", "_prio_num", "qty_35j"],
        ascending=[False, True, False, True, False],
    ).reset_index(drop=True)

    # ── 8. Allocation du stock réserve (logique approvisionneur) ─────────────
    # On parcourt les besoins dans l'ordre ci-dessus.
    # Pour chaque (barcode, taille), on décrémente le stock CENTRAL au fur et à mesure :
    # une pièce prise pour WEB n'est plus disponible pour les autres boutiques.
    #
    # Règle quantité minimum pour les tailles cœur :
    # Une taille cœur (part ≥ min_taille_share des ventes du barcode) se vend
    # immédiatement si on n'en envoie qu'une. On impose donc un minimum de
    # min_qty_core pièces pour ces tailles, dans la limite du stock disponible.
    _size_sales_alloc = (
        ventes_df.groupby(["barcode", "taille"])["qty"].sum().reset_index()
    )
    _barcode_total_alloc = ventes_df.groupby("barcode")["qty"].sum().reset_index()
    _size_sales_alloc = _size_sales_alloc.merge(_barcode_total_alloc, on="barcode", how="left", suffixes=("_t", "_b"))
    _size_sales_alloc["share"] = _size_sales_alloc["qty_t"] / _size_sales_alloc["qty_b"].replace(0, 1)
    core_sizes_alloc: dict[str, set] = (
        _size_sales_alloc[_size_sales_alloc["share"] >= min_taille_share]
        .groupby("barcode")["taille"].apply(set).to_dict()
    )

    reserve_pool = reserve_stock.set_index(["barcode", "taille"])["stock_reserve"].to_dict()

    proposed_qtys = []
    for _, row in needs.iterrows():
        key       = (row["barcode"], row["taille"])
        available = int(max(0, reserve_pool.get(key, 0)))
        qty       = int(min(max(0, row["besoin_theorique"]), available))

        # Pour les tailles cœur, on complète au moins jusqu'au plancher de stock
        # (min_stock_floor) — sans dépasser : la boutique atterrit pile au plancher
        # (ex. stock 1 → +1 = 2, pas +2 = 3), sauf si la couverture exige davantage.
        if qty > 0 and row["taille"] in core_sizes_alloc.get(row["barcode"], set()):
            besoin_plancher = max(0, min_stock_floor - int(row["stock"]))
            qty = max(qty, min(besoin_plancher, available))

        proposed_qtys.append(qty)
        if qty > 0:
            reserve_pool[key] = available - qty   # Décrémenter pour les boutiques suivantes

    needs = needs.copy()
    needs["qte_proposee"] = proposed_qtys

    # ── 9. Propositions finales ───────────────────────────────────────────────
    # Garder uniquement : transfert > 0 ET (urgence réelle OU bestseller OU plancher).
    # Plancher : une taille qui se vend (≥ floor_min_sales) et sous le minimum
    # (stock < min_stock_floor) doit être réassortie même si sa couverture > cible.
    _sous_plancher = (
        (needs["qty_35j"] >= floor_min_sales)
        & (needs["stock"] < min_stock_floor)
        & needs["_is_core"]
    )
    proposed = needs[
        (needs["qte_proposee"] > 0) &
        (
            (needs["couverture_jours"] < target_days) |
            (needs["qty_35j"] >= 3) |
            _sous_plancher
        )
    ].copy()

    # ── 9b. Filtre courbe de tailles rompue ───────────────────────────────────
    # Si pour un (boutique, barcode), le réassort ne couvre que des tailles fringe
    # (part < min_taille_share des ventes) ET qu'aucune taille cœur n'est disponible
    # en stock boutique → on exclut ces transferts (inutile d'envoyer XL/XXL seul).
    proposed, taille_seule_exclus = _filter_tailles_isolees(
        proposed, stock_df, ventes_df,
        reserve=reserve,
        min_taille_share=min_taille_share,
        taille_exemptions=taille_exemptions,
        min_covered_mens=min_covered_mens,
        min_covered_shoes=min_covered_shoes,
        min_qty_core=min_qty_core,
        reserve_pool=reserve_pool,
    )

    # ── 9c. Colonne "tailles après réassort" ──────────────────────────────────
    # Affiche toutes les tailles qui seront présentes après les transferts :
    # stock actuel boutique + tailles proposées en réassort (même si stock=0).
    # Format : "36(1) · 38(2*) · 39(2*)" — l'astérisque indique une taille en réa.
    if not proposed.empty:
        # Stock actuel boutique (hors CENTRAL)
        _bstock = stock_df[(stock_df["boutique"] != reserve) & (stock_df["stock"] > 0)]
        _bstock_idx = _bstock.set_index(["boutique", "barcode", "taille"])["stock"].to_dict()

        # Tailles proposées par (boutique, barcode)
        _prop_map: dict[tuple, dict[str, int]] = {}
        for _, row in proposed.iterrows():
            key = (row["boutique"], row["barcode"])
            _prop_map.setdefault(key, {})[row["taille"]] = int(row["qte_proposee"])

        # Reconstruit la carte combinée par (boutique, barcode)
        _after_map: dict[tuple, str] = {}
        all_keys = set(_prop_map.keys()) | {(r.boutique, r.barcode) for r in
                       _bstock[["boutique","barcode"]].drop_duplicates().itertuples(index=False)}
        for (bout, bc) in _prop_map.keys():
            tailles_apres: dict[str, str] = {}
            # 1. Stock actuel
            for (b2, bc2, t), q in _bstock_idx.items():
                if b2 == bout and bc2 == bc:
                    tailles_apres[t] = f"{t}({int(q)})"
            # 2. Tailles en réassort (avec astérisque si pas déjà en stock)
            for t, q in _prop_map.get((bout, bc), {}).items():
                if t in tailles_apres:
                    tailles_apres[t] = f"{t}({int(_bstock_idx.get((bout,bc,t),0))}+{q})"
                else:
                    tailles_apres[t] = f"{t}({q}*)"
            parts = sorted(tailles_apres.values(), key=lambda x: x.lower())
            _after_map[(bout, bc)] = " · ".join(parts) if parts else "—"

        proposed["tailles_stock_boutique"] = [
            _after_map.get((r.boutique, r.barcode), "—")
            for r in proposed.itertuples()
        ]
    else:
        proposed["tailles_stock_boutique"] = pd.Series(dtype=str)

    # ── 10. Risques sans réserve ──────────────────────────────────────────────
    # Articles urgents pour lesquels CENTRAL ne peut rien apporter
    risk = needs[
        (needs["besoin_theorique"] > 0) &
        (needs["qte_proposee"] == 0) &
        (needs["couverture_jours"] < 14)
    ].copy()

    # ── 11. Top ventes réseau ─────────────────────────────────────────────────
    top = sales_global.merge(reserve_stock, on=["barcode", "taille"], how="left")
    top["stock_reserve"] = top["stock_reserve"].fillna(0)
    top = top.sort_values("qty_global", ascending=False).head(300).reset_index(drop=True)

    return proposed, risk, top, promo_exclus, taille_seule_exclus


# ──────────────────────────────────────────────────────────────────────────────
# Export Excel
# ──────────────────────────────────────────────────────────────────────────────

def safe_sheet(name: str) -> str:
    return re.sub(r"[\\/*?:\[\]]", "_", str(name))[:31]


def write_df(ws, df: pd.DataFrame) -> None:
    """Écrit un DataFrame dans une feuille openpyxl via ws.append() (beaucoup plus rapide que cell-by-cell)."""
    ws.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws.append([
            None if (isinstance(val, float) and math.isnan(val)) else val
            for val in row
        ])


def style_ws(ws, header_color: str = "1F4E78") -> None:
    fill   = PatternFill("solid", fgColor=header_color)
    font_h = Font(color="FFFFFF", bold=True, size=10)
    thin   = Side(style="thin", color="BDD7EE")
    ws.freeze_panes = "A2"
    # Pas de ws.auto_filter ici : les Tables Excel gèrent déjà les filtres
    # (auto_filter + Table sur la même plage bloque l'ouverture du fichier)
    for cell in ws[1]:
        cell.fill      = fill
        cell.font      = font_h
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = Border(bottom=thin)
    # Pour les grandes feuilles (>2000 lignes), on saute l'alignement cellule-par-cellule
    # (trop lent) — la mise en page reste lisible grâce aux Tables Excel
    if ws.max_row <= 2000:
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top")
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        w = max((len(str(c.value)) for c in list(col_cells)[:200] if c.value is not None), default=8)
        ws.column_dimensions[letter].width = min(max(w + 2, 10), 42)


def add_table(ws, name: str) -> None:
    if ws.max_row < 2 or ws.max_column < 1:
        return
    safe = re.sub(r"[^A-Za-z0-9_]", "_", name)[:240]
    ref  = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    tab  = Table(displayName=safe, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True,
        showFirstColumn=False, showLastColumn=False, showColumnStripes=False,
    )
    ws.add_table(tab)


def color_priorities(ws, proposed: pd.DataFrame) -> None:
    """Colore les lignes P1 en rouge, P2 en orange."""
    if proposed.empty or ws.max_row < 2:
        return
    cols = list(proposed.columns)
    if "Priorité" not in cols:
        return
    p_col = get_column_letter(cols.index("Priorité") + 1)
    rng   = f"A2:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'LEFT(${p_col}2,2)="P1"'],
        fill=PatternFill("solid", fgColor="FFC7CE"),
    ))
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'LEFT(${p_col}2,2)="P2"'],
        fill=PatternFill("solid", fgColor="FFEB9C"),
    ))


COLS_PROPOSED = [
    "boutique", "barcode", "taille", "marque", "saison",
    "stock", "stock_reserve",
    "tailles_stock_boutique",
    "qty_35j", "ventes_jour", "couverture_jours",
    "besoin_theorique", "qte_proposee",
    "priorite", "commentaire", "qty_global",
    "urgence_boutique_j", "promo_pct",
]

COLS_PROMO = [
    "boutique", "barcode", "taille", "marque", "saison",
    "stock", "stock_reserve",
    "qty_35j", "couverture_jours", "besoin_theorique",
    "priorite", "promo_pct",
]

COLS_TAILLE_ISOLEE = [
    "boutique", "barcode", "taille", "marque", "saison",
    "stock", "stock_reserve",
    "qty_35j", "couverture_jours", "besoin_theorique",
    "priorite", "raison_exclusion",
]

LABELS = {
    "boutique":            "Boutique",
    "barcode":             "Barcode",
    "taille":              "Taille",
    "marque":              "Marque",
    "saison":              "Saison",
    "stock":               "Stock boutique",
    "stock_reserve":       "Stock CENTRAL",
    "tailles_stock_boutique": "Tailles en stock boutique",
    "qty_35j":             "Ventes 35j",
    "ventes_jour":         "Ventes/jour",
    "couverture_jours":    "Couverture (j)",
    "besoin_theorique":    "Besoin théorique",
    "qte_proposee":        "Qté proposée",
    "priorite":            "Priorité",
    "commentaire":         "Commentaire",
    "qty_global":          "Ventes réseau 35j",
    "urgence_boutique_j":  "Urgence boutique (j)",
    "promo_pct":           "Remise moy. (%)",
    "raison_exclusion":    "Raison d'exclusion",
}


def df_for_export(df: pd.DataFrame, cols: list[str], drop_cols: list[str] | None = None) -> pd.DataFrame:
    """Sélectionne les colonnes disponibles et renomme avec les labels lisibles."""
    available = [c for c in cols if c in df.columns]
    out = df[available].copy()
    if drop_cols:
        out = out.drop(columns=[c for c in drop_cols if c in out.columns], errors="ignore")
    out.columns = [LABELS.get(c, c) for c in out.columns]
    return out


def build_excel(
    output:              Path,
    proposed:            pd.DataFrame,
    risk:                pd.DataFrame,
    top:                 pd.DataFrame,
    promo_exclus:        pd.DataFrame,
    taille_seule_exclus: pd.DataFrame,
    reserve:             str,
    sales_start:         str,
    sales_end:           str,
    target_days:         int,
    sales_days:          int,
) -> None:

    wb = Workbook()
    boutiques = sorted(proposed["boutique"].unique()) if not proposed.empty else []

    # ── Feuille Synthèse ──────────────────────────────────────────────────────
    ws0 = wb.active
    ws0.title = "Synthèse"

    now = datetime.now().strftime("%d/%m/%Y %H:%M")
    periode = f"{sales_start} → {sales_end}" if sales_start else f"{sales_days} jours"
    total_qte = int(proposed["qte_proposee"].sum()) if not proposed.empty else 0

    prio_summary = (
        proposed.groupby("priorite")["qte_proposee"].agg(["count", "sum"])
                .reset_index()
                .sort_values("priorite")
        if not proposed.empty else pd.DataFrame(columns=["priorite", "count", "sum"])
    )

    rows_synth = [
        ["AGENT RÉASSORT MULTI-BOUTIQUES", ""],
        ["Généré le",            now],
        ["Réserve utilisée",     reserve],
        ["Période ventes",       periode],
        ["Objectif couverture",  f"{target_days} jours"],
        ["", ""],
        ["RÉSUMÉ GLOBAL", ""],
        ["Boutiques analysées",          len(boutiques)],
        ["Lignes de transfert proposées", len(proposed)],
        ["Quantité totale à transférer", total_qte],
        ["Références en risque (sans réserve)", len(risk)],
        ["Exclus courbe tailles rompue", len(taille_seule_exclus)],
        ["", ""],
        ["PRIORITÉ", "Qtés proposées"],
    ]
    for _, r in prio_summary.iterrows():
        rows_synth.append([r["priorite"], int(r["sum"])])

    rows_synth += [["", ""], ["BOUTIQUE", "Qtés à recevoir"]]
    if not proposed.empty:
        by_b = proposed.groupby("boutique")["qte_proposee"].sum().sort_values(ascending=False)
        for b, q in by_b.items():
            rows_synth.append([b, int(q)])

    # Ordre d'allocation : WEB en tête, puis boutiques par urgence top30
    rows_synth += [["", ""], ["ORDRE D'ALLOCATION (comment CENTRAL a été distribué)", ""]]
    rows_synth += [["Rang", "Boutique", "Cov. min top30 (j)", "Vol. ventes top30", "Qtés reçues"]]
    if not proposed.empty:
        # Reconstruction de l'ordre depuis les colonnes internes
        alloc_order = (
            proposed[["boutique", "_boutique_urgency", "_boutique_volume", "qte_proposee"]]
            .groupby("boutique")
            .agg(
                urgency=("_boutique_urgency", "first"),
                volume=("_boutique_volume", "first"),
                qte=("qte_proposee", "sum"),
            )
            .reset_index()
            .sort_values(["urgency", "volume"], ascending=[True, False])
        )
        for rank, row in enumerate(alloc_order.itertuples(), 1):
            label = " ◀ PRIORITÉ 1 (WEB)" if row.boutique == "WEB" else ""
            cov = 0.0 if row.urgency <= 0 else round(row.urgency, 1)
            rows_synth.append([rank, f"{row.boutique}{label}", cov, int(row.volume), int(row.qte)])

    for row in rows_synth:
        ws0.append(row)

    ws0["A1"].font  = Font(bold=True, size=14, color="1F4E78")
    ws0["A7"].font  = Font(bold=True, size=11)
    ws0["A13"].font = Font(bold=True, size=11)
    # Trouver la ligne "BOUTIQUE"
    boutique_row = 13 + 1 + len(prio_summary) + 2
    ws0.cell(boutique_row, 1).font = Font(bold=True, size=11)
    # Trouver la ligne "ORDRE D'ALLOCATION"
    ordre_row = boutique_row + 1 + (len(proposed["boutique"].unique()) if not proposed.empty else 0) + 2
    ws0.cell(ordre_row, 1).font = Font(bold=True, size=11, color="1F4E78")
    ws0.column_dimensions["A"].width = 8
    ws0.column_dimensions["B"].width = 36
    ws0.column_dimensions["C"].width = 22
    ws0.column_dimensions["D"].width = 22
    ws0.column_dimensions["E"].width = 16

    # ── Tous les transferts ────────────────────────────────────────────────────
    ws1 = wb.create_sheet("Tous transferts")
    if not proposed.empty:
        df_out = df_for_export(proposed, COLS_PROPOSED)
        df_out = df_out.sort_values(["Priorité", "Ventes réseau 35j"], ascending=[True, False])
        write_df(ws1, df_out)
        style_ws(ws1, "1F4E78")
        add_table(ws1, "tbl_tous_transferts")
        color_priorities(ws1, df_out)

    # ── Une feuille par boutique ───────────────────────────────────────────────
    for boutique in boutiques:
        df_b = proposed[proposed["boutique"] == boutique].copy()
        if df_b.empty:
            continue
        ws_b  = wb.create_sheet(safe_sheet(boutique))
        df_out = df_for_export(df_b, COLS_PROPOSED, drop_cols=["boutique"])
        df_out = df_out.sort_values("Priorité")
        write_df(ws_b, df_out)
        style_ws(ws_b, "2E4057")
        add_table(ws_b, f"tbl_{re.sub(r'[^A-Za-z0-9]', '_', boutique)}")
        color_priorities(ws_b, df_out)

    # ── Alertes promo ─────────────────────────────────────────────────────────
    ws_promo = wb.create_sheet("Alertes promo >20%")
    if not promo_exclus.empty:
        cols_p = [c for c in COLS_PROMO if c in promo_exclus.columns]
        df_out = promo_exclus[cols_p].copy()
        df_out.columns = [LABELS.get(c, c) for c in df_out.columns]
        # Formater la remise en %
        if "Remise moy. (%)" in df_out.columns:
            df_out["Remise moy. (%)"] = (df_out["Remise moy. (%)"] * 100).round(1)
        df_out = df_out.sort_values("Remise moy. (%)", ascending=False)
        write_df(ws_promo, df_out)
        style_ws(ws_promo, "7B2C2C")
        add_table(ws_promo, "tbl_alertes_promo")
        # Colorer les remises extrêmes (>50%) en rouge foncé
        if "Remise moy. (%)" in df_out.columns:
            p_col = get_column_letter(list(df_out.columns).index("Remise moy. (%)") + 1)
            rng = f"A2:{get_column_letter(ws_promo.max_column)}{ws_promo.max_row}"
            ws_promo.conditional_formatting.add(rng, FormulaRule(
                formula=[f"${p_col}2>=50"],
                fill=PatternFill("solid", fgColor="FF0000"),
            ))

    # ── Tailles isolées (courbe rompue) ───────────────────────────────────────
    ws_taille = wb.create_sheet("Tailles isolées")
    if not taille_seule_exclus.empty:
        cols_t = [c for c in COLS_TAILLE_ISOLEE if c in taille_seule_exclus.columns]
        df_out = taille_seule_exclus[cols_t].copy()
        df_out.columns = [LABELS.get(c, c) for c in df_out.columns]
        df_out = df_out.sort_values(["Boutique", "Barcode"])
        write_df(ws_taille, df_out)
        style_ws(ws_taille, "7B4A1A")
        add_table(ws_taille, "tbl_tailles_isolees")
    else:
        ws_taille.cell(1, 1, "Aucune courbe de tailles rompue détectée.")

    # ── Risques sans réserve ───────────────────────────────────────────────────
    ws_risk = wb.create_sheet("Risques sans réserve")
    if not risk.empty:
        cols_risk = ["boutique", "barcode", "taille", "marque", "stock",
                     "qty_35j", "ventes_jour", "couverture_jours",
                     "besoin_theorique", "priorite", "commentaire"]
        df_out = df_for_export(risk, cols_risk)
        df_out = df_out.sort_values("Couverture (j)")
        write_df(ws_risk, df_out)
        style_ws(ws_risk, "C00000")
        add_table(ws_risk, "tbl_risques")

    # ── Top ventes réseau ──────────────────────────────────────────────────────
    ws_top = wb.create_sheet("Top ventes réseau")
    if not top.empty:
        cols_top = ["barcode", "taille", "qty_global", "stock_reserve"]
        lbl_top  = {
            "barcode":       "Barcode",
            "taille":        "Taille",
            "qty_global":    "Ventes réseau 35j",
            "stock_reserve": f"Stock {reserve}",
        }
        df_out = top[[c for c in cols_top if c in top.columns]].copy()
        df_out.columns = [lbl_top.get(c, c) for c in df_out.columns]
        write_df(ws_top, df_out)
        style_ws(ws_top, "375623")
        add_table(ws_top, "tbl_top_ventes")

    # ── Paramètres ────────────────────────────────────────────────────────────
    ws_p = wb.create_sheet("Paramètres")
    for row in [
        ["Paramètre",         "Valeur",      "Description"],
        ["sales_days",        sales_days,    "Période des ventes (jours)"],
        ["target_days",       target_days,   "Couverture cible (jours)"],
        ["reserve",           reserve,       "Boutique réserve"],
        ["p1_seuil_jours",    7,             "Seuil P1 : rupture imminente si < x jours"],
        ["p2_seuil_jours",    14,            "Seuil P2 : à surveiller si < x jours"],
        ["p3_min_ventes_35j", 3,             "Ventes min 35j pour inclure même si bien couvert"],
        ["min_taille_share",  0.15,          "Part min de ventes pour qu'une taille soit 'cœur' (filtre courbe rompue)"],
        ["min_qty_core",      2,             "Qté min à transférer pour les tailles cœur (évite rupture immédiate)"],
        ["min_covered_mens",  3,             "Nb tailles homme (XS/S/M/L) à couvrir min pour valider un transfert"],
        ["mens_sizes",        "XS,S,M,L",   "Tailles considérées comme 'homme' pour la règle courbe"],
        ["min_covered_shoes", 3,             "Nb pointures (41–45 incl. demi-tailles) à couvrir min pour valider un transfert chaussures"],
        ["shoes_core_range",  "42.0–44.0",   "Plage de pointures cœur chaussures (incl. 42.5, 43.5…)"],
    ]:
        ws_p.append(row)
    style_ws(ws_p, "404040")

    wb.save(output)


# ──────────────────────────────────────────────────────────────────────────────
# Sélecteur de fichiers (GUI Tkinter)
# ──────────────────────────────────────────────────────────────────────────────

def update_stock_central(src) -> bool:
    """Charge un export Fastmag et le copie comme stock_central.xls (réf. du pipeline).

    - ne garde que les lignes du magasin CENTRAL
    - sauvegarde l'ancien stock_central.xls (backup horodaté)
    - écrit le nouveau au même format (tab, latin-1, CRLF)
    Renvoie True si la mise à jour a réussi.
    """
    import shutil
    src = Path(src)
    dest = Path(__file__).parent / "stock_central.xls"
    cols_requises = ["Magasin", "Stock", "gencod"]
    try:
        with open(src, "r", encoding="latin-1", errors="replace", newline="") as f:
            contenu = f.read()
    except Exception as exc:
        print(f"    ⚠️  Export CENTRAL illisible ({exc}) — stock CENTRAL inchangé")
        return False

    contenu = contenu.replace("\r\n", "\n").replace("\r", "\n")
    lignes = [l for l in contenu.split("\n")]
    while lignes and lignes[-1] == "":
        lignes.pop()
    if not lignes:
        print("    ⚠️  Export CENTRAL vide — stock CENTRAL inchangé")
        return False

    entete = lignes[0].split("\t")
    has_ref = any(c.lower() in ("référence", "reference", "réf") for c in entete)
    manquantes = [c for c in cols_requises if c not in entete] + ([] if has_ref else ["Référence"])
    if manquantes:
        print(f"    ⚠️  Pas un export Fastmag (colonnes manquantes : {', '.join(manquantes)}) — stock CENTRAL inchangé")
        return False

    idx_mag   = entete.index("Magasin")
    idx_stock = entete.index("Stock")
    gardees = [lignes[0]]
    total = 0.0
    for l in lignes[1:]:
        if not l.strip():
            continue
        champs = l.split("\t")
        if len(champs) <= idx_mag:
            continue
        if champs[idx_mag].strip().upper() in ("CENTRAL", "PRESTA"):
            gardees.append(l)
            try:
                total += float(champs[idx_stock].replace(",", ".").replace(" ", "") or 0)
            except (ValueError, IndexError):
                pass

    if len(gardees) <= 1:
        print("    ⚠️  Aucune ligne CENTRAL dans l'export — stock CENTRAL inchangé")
        return False

    if dest.exists():
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        try:
            shutil.copy2(dest, dest.with_name(f"stock_central_backup_{stamp}.xls"))
        except Exception:
            pass

    with open(dest, "w", encoding="latin-1", errors="replace", newline="") as f:
        f.write("\r\n".join(gardees) + "\r\n")

    print(f"    Export CENTRAL chargé : {len(gardees)-1:,} réfs  |  {int(total):,} unités → stock_central.xls ✓")
    return True


def pick_files_gui(default_reserve: str) -> dict:
    try:
        import tkinter as tk
        from tkinter import filedialog, messagebox, simpledialog
    except Exception as exc:
        raise RuntimeError(
            "Tkinter non disponible. Utilise --stock et --ventes en ligne de commande."
        ) from exc

    root = tk.Tk()
    root.withdraw()
    root.update()
    ft = [("CSV / Excel", "*.csv *.xls *.xlsx *.xlsm"), ("Tous", "*.*")]

    try:
        messagebox.showinfo(
            "Agent Réassort Multi-Boutiques",
            "Sélectionne les fichiers dans l'ordre :\n\n"
            "1. Export STOCK CENTRAL Fastmag (optionnel — pour mettre à jour)\n"
            "2. Fichier STOCK (toutes boutiques)\n"
            "3. Fichier VENTES (35 derniers jours)\n"
            "4. Emplacement du fichier Excel de sortie"
        )

        # ── 1/4 — Export stock CENTRAL (optionnel) ────────────────────────
        central_p = None
        maj_central = messagebox.askyesno(
            "Stock CENTRAL",
            "Veux-tu charger un nouvel export stock CENTRAL Fastmag ?\n\n"
            "• Oui  → choisis l'export, il remplacera stock_central.xls\n"
            "• Non  → on garde le stock CENTRAL déjà chargé"
        )
        if maj_central:
            central_p = filedialog.askopenfilename(
                title="1/4 — Export STOCK CENTRAL Fastmag", filetypes=ft
            )
            if not central_p:
                central_p = None  # annulation = on garde l'existant

        stock_p = filedialog.askopenfilename(title="2/4 — Fichier STOCK", filetypes=ft)
        if not stock_p: raise SystemExit("Annulé.")

        ventes_p = filedialog.askopenfilename(title="3/4 — Fichier VENTES", filetypes=ft)
        if not ventes_p: raise SystemExit("Annulé.")

        reserve_name = simpledialog.askstring(
            "Réserve", "Nom de la boutique réserve :", initialvalue=default_reserve
        ) or default_reserve

        output_p = filedialog.asksaveasfilename(
            title="4/4 — Fichier Excel de sortie",
            defaultextension=".xlsx",
            initialfile=f"reassort_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            filetypes=[("Excel", "*.xlsx")],
        )
        if not output_p: raise SystemExit("Annulé.")

        return {
            "stock":   Path(stock_p),
            "ventes":  Path(ventes_p),
            "output":  Path(output_p),
            "reserve": reserve_name,
            "central":  Path(central_p) if central_p else None,
        }
    finally:
        root.destroy()


# ──────────────────────────────────────────────────────────────────────────────
# CLI
# ──────────────────────────────────────────────────────────────────────────────

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Agent Réassort Multi-Boutiques — génère un plan de transfert depuis la réserve.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    p.add_argument("--stock",        default=None, help="Fichier stock (CSV/Excel)")
    p.add_argument("--ventes",       default=None, help="Fichier ventes (CSV/Excel)")
    p.add_argument("--output",       default=None, help="Fichier Excel de sortie")
    p.add_argument("--reserve",      default=DEFAULT_RESERVE,     help=f"Boutique réserve (défaut : {DEFAULT_RESERVE})")
    p.add_argument("--target-days",  default=DEFAULT_TARGET_DAYS, type=int, help=f"Couverture cible en jours (défaut : {DEFAULT_TARGET_DAYS})")
    p.add_argument("--sales-days",   default=DEFAULT_SALES_DAYS,  type=int, help=f"Période ventes si pas de dates (défaut : {DEFAULT_SALES_DAYS})")
    p.add_argument("--promo-seuil",       default=0.20,  type=float, help="Seuil de remise pour exclusion (défaut : 0.20 = 20%%)")
    p.add_argument("--promo-exemptions", default=None,  type=str,   help="Boutiques exemptées du filtre promo, séparées par virgule (défaut : WEB)")
    p.add_argument("--min-taille-share",   default=0.15, type=float, help="Part min de ventes pour qu'une taille soit 'cœur' (défaut : 0.15 = 15%%)")
    p.add_argument("--taille-exemptions", default=None, type=str,   help="Boutiques exemptées du filtre courbe tailles, séparées par virgule (défaut : WEB)")
    p.add_argument("--min-qty-core",      default=2,    type=int,   help="Quantité minimum à transférer pour les tailles cœur (défaut : 2)")
    p.add_argument("--min-covered-mens",  default=3,    type=int,   help="Nb de tailles cœur (XS/S/M/L) à avoir couvertes pour valider le transfert (défaut : 3)")
    p.add_argument("--min-covered-shoes", default=3,    type=int,   help="Nb de pointures cœur (41-45) à avoir couvertes pour valider le transfert chaussures (défaut : 3)")
    p.add_argument("--min-saison-year",  default=26,   type=int,   help="Année saison minimum (2 chiffres) à inclure dans le réassort (défaut : 26 = 2026). 0 = désactiver.")
    p.add_argument("--min-stock-floor",  default=MIN_STOCK_FLOOR,       type=int, help=f"Stock minimum garanti par taille qui se vend (défaut : {MIN_STOCK_FLOOR})")
    p.add_argument("--floor-min-sales",  default=MIN_STOCK_FLOOR_SALES, type=int, help=f"Ventes 35j à partir desquelles le plancher s'applique (défaut : {MIN_STOCK_FLOOR_SALES})")
    p.add_argument("--stock-central", default=None, help="Export Fastmag stock CENTRAL à charger (met à jour stock_central.xls)")
    p.add_argument("--no-gui",       action="store_true", help="Désactiver le sélecteur de fichiers")
    return p.parse_args()


# ──────────────────────────────────────────────────────────────────────────────
# Email recap
# ──────────────────────────────────────────────────────────────────────────────

EMAIL_TO   = "anthony.poignard@footkorner.com"
EMAIL_FROM = "anthony.poignard@footkorner.com"

_EMAIL_CSS = """
body{font-family:Arial,sans-serif;background:#f4f6f9;margin:0;padding:20px;color:#222}
.wrapper{max-width:820px;margin:0 auto;background:#fff;border-radius:8px;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,.12)}
.header{background:#1F4E78;color:#fff;padding:28px 32px 20px}
.header h1{margin:0 0 6px;font-size:22px;letter-spacing:.5px}
.header p{margin:0;opacity:.85;font-size:13px}
.body{padding:28px 32px}
.kpi-row{display:flex;gap:16px;margin-bottom:28px;flex-wrap:wrap}
.kpi{flex:1;min-width:130px;background:#EBF2FA;border-left:4px solid #1F4E78;border-radius:4px;padding:14px 16px}
.kpi .num{font-size:28px;font-weight:700;color:#1F4E78;line-height:1}
.kpi .lbl{font-size:11px;color:#555;margin-top:4px;text-transform:uppercase;letter-spacing:.5px}
.kpi.warn{background:#FFF3CD;border-color:#FFC107}.kpi.warn .num{color:#856404}
.kpi.danger{background:#FDE8E8;border-color:#C00000}.kpi.danger .num{color:#C00000}
.section-title{font-size:15px;font-weight:700;color:#1F4E78;border-bottom:2px solid #BDD7EE;padding-bottom:6px;margin:24px 0 12px}
table{width:100%;border-collapse:collapse;font-size:13px}
th{background:#1F4E78;color:#fff;padding:8px 10px;text-align:left;font-weight:600;font-size:12px}
td{padding:7px 10px;border-bottom:1px solid #E8EEF4;vertical-align:top}
tr:nth-child(even) td{background:#F7FAFD}
.badge{display:inline-block;padding:2px 7px;border-radius:10px;font-size:11px;font-weight:700;white-space:nowrap}
.p1r{background:#FFC7CE;color:#9C0006}.p1j{background:#FFAB91;color:#BF360C}
.p2{background:#FFEB9C;color:#7D6608}.p3{background:#C6EFCE;color:#276221}
.p4{background:#DEEBF7;color:#1F4E78}
.boutique-block{margin-bottom:20px;border:1px solid #D9E6F2;border-radius:6px;overflow:hidden}
.boutique-header{background:#2E4057;color:#fff;padding:9px 14px;display:flex;justify-content:space-between;align-items:center}
.boutique-header .name{font-weight:700;font-size:14px}
.boutique-header .meta{font-size:12px;opacity:.85}
.boutique-header.web{background:#1F4E78}
.boutique-inner table{margin:0}.boutique-inner td{font-size:12px}
.footer{background:#F0F4F8;padding:16px 32px;font-size:11px;color:#777;border-top:1px solid #DCE4EE}
.info-box{background:#D1ECF1;border:1px solid #BEE5EB;border-radius:5px;padding:12px 16px;margin-bottom:20px;font-size:13px;color:#0C5460}
.cmd-table td{font-size:13px;padding:8px 12px}.cmd-table tr:hover td{background:#EBF2FA}
"""

_BADGE = {
    "P1 - Rupture":   "<span class='badge p1r'>P1 – Rupture</span>",
    "P1 - <7 jours":  "<span class='badge p1j'>P1 – &lt;7 jours</span>",
    "P2 - <14 jours": "<span class='badge p2'>P2 – &lt;14 jours</span>",
    "P3 - Best-seller":"<span class='badge p3'>P3 – Best-seller</span>",
    "P4 - Opportunite":"<span class='badge p4'>P4 – Opportunité</span>",
}


def _badge(prio: str) -> str:
    for k, v in _BADGE.items():
        if k.lower() in str(prio).lower():
            return v
    return f"<span class='badge p4'>{prio}</span>"


def build_email_html(
    proposed:            pd.DataFrame,
    risk:                pd.DataFrame,
    promo_exclus:        pd.DataFrame,
    taille_seule_exclus: pd.DataFrame,
    sales_start:         str,
    sales_end:           str,
    target_days:         int,
    sales_days:          int,
    reserve:             str,
    run_date:            datetime | None = None,
    max_lines_per_boutique: int = 60,
) -> str:
    """Génère le HTML du mail récap réassort."""
    if run_date is None:
        run_date = datetime.now()
    ts = run_date.strftime("%d/%m/%Y %H:%M")

    total_qte    = int(proposed["qte_proposee"].astype(float).sum()) if not proposed.empty else 0
    total_lignes = len(proposed)
    nb_boutiques = proposed["boutique"].nunique() if not proposed.empty else 0
    nb_p1        = int(proposed[proposed["priorite"].str.contains("P1", na=False)]["qte_proposee"].astype(float).sum()) if not proposed.empty else 0
    nb_risques   = len(risk) if not risk.empty else 0

    # ── Commandes à préparer ────────────────────────────────────────────
    cmd_rows = ""
    if not proposed.empty:
        cmd = (
            proposed.groupby("boutique")
            .agg(nb_lignes=("qte_proposee", "count"), nb_pcs=("qte_proposee", lambda x: int(x.astype(float).sum())))
            .sort_values("nb_pcs", ascending=False)
        )
        for bout, row in cmd.iterrows():
            icon = "🌐 " if str(bout).upper() == "WEB" else ""
            cmd_rows += f"<tr><td><strong>{icon}{bout}</strong></td><td style='text-align:center'>{row['nb_lignes']}</td><td style='text-align:center'><strong>{row['nb_pcs']} pcs</strong></td></tr>\n"

    # ── Priorités ────────────────────────────────────────────────────────
    prio_rows = ""
    if not proposed.empty:
        for prio, grp in proposed.groupby("priorite"):
            qte = int(grp["qte_proposee"].astype(float).sum())
            prio_rows += f"<tr><td>{_badge(prio)}</td><td>{len(grp)}</td><td><strong>{qte} pcs</strong></td></tr>\n"

    # ── Info filtres ─────────────────────────────────────────────────────
    n_promo  = len(promo_exclus)  if not promo_exclus.empty  else 0
    n_taille = len(taille_seule_exclus) if not taille_seule_exclus.empty else 0
    info_txt = (
        f"✅ Filtres : <strong>remise &gt; 20%</strong> → {n_promo} besoins exclus"
        f" &nbsp;|&nbsp; <strong>courbe tailles rompue</strong> → {n_taille} transferts exclus"
        " &nbsp;|&nbsp; WEB exempté"
    )

    # Le détail par boutique n'est plus dans l'email : il est fourni en PDF individuel.

    html = f"""<!DOCTYPE html><html lang="fr"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<style>{_EMAIL_CSS}</style></head><body>
<div class="wrapper">
<div class="header">
  <h1>📦 Plan de Réassort Multi-Boutiques</h1>
  <p>Période de référence : {sales_start} → {sales_end} ({sales_days} jours) &nbsp;|&nbsp; Généré le {ts} &nbsp;|&nbsp; Réserve : {reserve} &nbsp;|&nbsp; Cible : {target_days} jours</p>
</div>
<div class="body">

<div class="section-title">🗂️ Commandes à préparer pour {reserve}</div>
<table class="cmd-table">
  <tr><th>Magasin</th><th style="text-align:center">Nb lignes</th><th style="text-align:center">Nb pièces</th></tr>
  {cmd_rows}
</table>

<div class="section-title">Synthèse globale</div>
<div class="kpi-row">
  <div class="kpi"><div class="num">{total_qte:,}</div><div class="lbl">Pièces à transférer</div></div>
  <div class="kpi"><div class="num">{total_lignes:,}</div><div class="lbl">Lignes de transfert</div></div>
  <div class="kpi"><div class="num">{nb_boutiques}</div><div class="lbl">Boutiques</div></div>
  <div class="kpi danger"><div class="num">{nb_p1:,}</div><div class="lbl">Urgences P1</div></div>
  <div class="kpi warn"><div class="num">{nb_risques:,}</div><div class="lbl">Risques sans réserve</div></div>
</div>

<div class="info-box">{info_txt}</div>

<div class="section-title">Répartition par priorité</div>
<table>
  <tr><th>Priorité</th><th>Nb références</th><th>Quantité</th></tr>
  {prio_rows}
</table>

<div class="info-box">📄 Le détail par boutique est en pièce jointe : un PDF par magasin (fichier zip).</div>

</div>
<div class="footer">Agent Réassort Footkorner &nbsp;·&nbsp; {ts} &nbsp;|&nbsp; Pièces jointes : Excel réassort + Import Fastmag + PDF par magasin (zip)</div>
</div></body></html>"""

    return html


def send_email_recap(
    html_body:     str,
    excel_path:    Path,
    fastmag_path:  Path | None,
    script_dir:    Path,
    run_date:      datetime | None = None,
    pdf_zip_path:  Path | None = None,
) -> bool:
    """
    Envoie le récap réassort par email via Gmail API (OAuth2).

    Fichiers requis dans script_dir (dossier du script) :
      - credentials.json   (client OAuth2)
      - token.json         (token généré au premier lancement)
    """
    import base64
    import json
    from email.mime.multipart import MIMEMultipart
    from email.mime.text      import MIMEText
    from email.mime.base      import MIMEBase
    from email               import encoders

    try:
        from google.oauth2.credentials          import Credentials
        from google_auth_oauthlib.flow          import InstalledAppFlow
        from google.auth.transport.requests     import Request
        from googleapiclient.discovery          import build
    except ImportError:
        print(
            "\n⚠️  Email non envoyé — librairies Google manquantes.\n"
            "   Lance : pip install google-auth google-auth-oauthlib google-api-python-client"
        )
        return False

    if run_date is None:
        run_date = datetime.now()

    SCOPES = [
        "https://www.googleapis.com/auth/gmail.modify",
        "https://www.googleapis.com/auth/drive",
    ]
    creds_path = script_dir / "credentials.json"
    token_path  = script_dir / "token.json"

    if not creds_path.exists():
        print(
            "\n⚠️  Email non envoyé — credentials.json introuvable.\n"
            f"   Place credentials.json et token.json dans :\n"
            f"   {script_dir}"
        )
        return False

    # ── Chargement / rafraîchissement du token ───────────────────────────
    creds = None
    if token_path.exists():
        creds = Credentials.from_authorized_user_file(str(token_path), SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(str(creds_path), SCOPES)
            creds = flow.run_local_server(port=0)
        token_path.write_text(creds.to_json(), encoding="utf-8")

    # ── Construction du message ──────────────────────────────────────────
    subject = f"Plan de Réassort Footkorner — {run_date.strftime('%d/%m/%Y')}"
    msg = MIMEMultipart("mixed")
    msg["From"]    = EMAIL_FROM
    msg["To"]      = EMAIL_TO
    msg["Subject"] = subject

    msg.attach(MIMEText(html_body, "html", "utf-8"))

    def _attach(path: Path, mime_type: str = "application/octet-stream") -> None:
        if path and path.exists():
            with open(path, "rb") as f:
                part = MIMEBase(*mime_type.split("/"))
                part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", "attachment", filename=path.name)
            msg.attach(part)

    _attach(excel_path, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    if fastmag_path:
        _attach(fastmag_path, "text/plain")
    if pdf_zip_path:
        _attach(pdf_zip_path, "application/zip")

    # ── Envoi via Gmail API ──────────────────────────────────────────────
    try:
        service = build("gmail", "v1", credentials=creds)
        raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()
        service.users().messages().send(userId="me", body={"raw": raw}).execute()
        print(f"✅  Email envoyé → {EMAIL_TO}")
        return True
    except Exception as e:
        print(f"⚠️  Échec envoi email : {e}")
        return False


# ──────────────────────────────────────────────────────────────────────────────
# Génération fichier d'import Fastmag
# ──────────────────────────────────────────────────────────────────────────────

def generate_fastmag_import(
    proposed:      pd.DataFrame,
    output_path:   Path,
    script_dir:    Path,
    run_date:      datetime | None = None,
    delivery_days: int = 2,
) -> tuple[int, int, list[str]]:
    """
    Génère le fichier d'import Fastmag (.txt tab-séparé).

    Format : 1 ligne numérotée par (boutique × barcode × taille × qté > 0).
    Les colonnes fixes (Commande=BL, devise=EUR, …) sont remplies automatiquement.

    Sources de données (dans script_dir) :
      - fastmag_bdd.xls  ou  fastmag_bdd.xls.txt → Designation + Couleur
      - prix de gros.csv                          → PU (prix unitaire HT)
      - Footkorner - Listing mag.xlsx             → code mag → NUM Fastmag

    Retourne : (nb_lignes_export, nb_boutiques_exportées, boutiques_sans_num)
    """
    from datetime import timedelta

    if run_date is None:
        run_date = datetime.now()

    date_str   = run_date.strftime("%d/%m/%Y")
    deliv_date = (run_date + timedelta(days=delivery_days)).strftime("%d/%m/%Y")
    ref_batch  = f"REASUCC{run_date.strftime('%d%m%Y')}"

    # ── Listing mag : code_mag → NUM ────────────────────────────────────
    listing_path = script_dir / "Footkorner - Listing mag.xlsx"
    if not listing_path.exists():
        print("⚠️  Listing mag introuvable — génération Fastmag annulée")
        return 0, 0, []

    listing_df = pd.read_excel(listing_path, dtype=str)
    mag_to_num: dict[str, int] = {}
    for _, lrow in listing_df.iterrows():
        num = str(lrow.get("NUM", "")).strip()
        if not num:
            continue
        try:
            num_int = int(float(num))
        except ValueError:
            continue
        # Indexer par code mag fastmag ET par nom MAG (les deux)
        for col in ("code mag fastmag", "MAG", "RAISON SOCIALE"):
            val = str(lrow.get(col, "")).strip().upper()
            if val:
                mag_to_num.setdefault(val, num_int)

    # ── Overrides manuels ─────────────────────────────────────────────────
    # FK2 = Toulouse → magasin n°4
    mag_to_num.setdefault("FK2", 4)
    mag_to_num.setdefault("TOULOUSE", 4)
    # WEB = compte web Fastmag → magasin n°164
    mag_to_num.setdefault("WEB", 164)

    # ── Couleur CENTRAL : (barcode, taille) → couleur avec le plus de stock ─
    # Source prioritaire pour la couleur : si plusieurs couleurs existent dans
    # la BDD pour le même (ref, taille), on prend celle qui a du stock en CENTRAL.
    central_couleur: dict[tuple[str, str], str] = {}
    _central_coul_path = script_dir / "stock_central.xls"
    if _central_coul_path.exists():
        try:
            _df_pc = pd.read_csv(_central_coul_path, sep='\t', encoding='latin1', dtype=str)
            _df_pc.columns = [c.strip() for c in _df_pc.columns]
            _ref_c  = next(c for c in _df_pc.columns if c.lower() in ('référence','reference','réf'))
            _df_pc["_bc"]  = _df_pc[_ref_c].str.strip()
            _df_pc["_ta"]  = _df_pc["Taille"].str.strip()
            _df_pc["_co"]  = _df_pc["Couleur"].str.strip()
            _df_pc["_st"]  = pd.to_numeric(
                _df_pc["Stock"].str.replace(',', '.', regex=False), errors='coerce'
            ).fillna(0)
            # Pour chaque (barcode, taille), garde la couleur avec le plus de stock
            _df_pc_s = _df_pc.sort_values("_st", ascending=False)
            for _, r in _df_pc_s.iterrows():
                key = (r["_bc"], r["_ta"])
                if key not in central_couleur and r["_co"]:
                    central_couleur[key] = r["_co"]
        except Exception as _exc:
            pass  # fallback BDD utilisé silencieusement

    # ── BDD Fastmag : (ref_clean, taille_clean) → (Designation, Couleur) ─
    bdd_path = script_dir / "fastmag_bdd.xls"
    if not bdd_path.exists():
        bdd_path = script_dir / "fastmag_bdd.xls.txt"

    bdd_lookup:    dict[tuple[str, str], tuple[str, str]] = {}
    bdd_ref_only:  dict[str, tuple[str, str]]             = {}

    if bdd_path.exists():
        try:
            bdd_df = pd.read_csv(
                bdd_path, sep="\t", dtype=str, low_memory=False, encoding="latin1"
            )
            bdd_df["_ref"]    = bdd_df["Reference_Article"].fillna("").str.strip().str.lstrip("% .")
            bdd_df["_taille"] = bdd_df["Taille"].fillna("").str.strip()
            for _, brow in bdd_df.iterrows():
                ref    = brow["_ref"]
                taille = brow["_taille"]
                desig  = str(brow.get("Designation", "")).strip()
                coul   = str(brow.get("Couleur",     "")).strip()
                key    = (ref, taille)
                if key not in bdd_lookup:
                    bdd_lookup[key] = (desig, coul)
                if ref and ref not in bdd_ref_only:
                    bdd_ref_only[ref] = (desig, coul)
            print(f"    BDD Fastmag : {len(bdd_lookup):,} couples (ref, taille) indexés")
        except Exception as exc:
            print(f"⚠️  BDD Fastmag illisible ({exc}) — Designation/Couleur vides")
    else:
        print("⚠️  BDD Fastmag absente — Designation/Couleur vides")

    # ── Prix de gros : ref → float ───────────────────────────────────────
    prix_path   = script_dir / "prix de gros.csv"
    prix_lookup: dict[str, float] = {}

    if prix_path.exists():
        try:
            prix_df = pd.read_csv(prix_path, sep=";", dtype=str, encoding="utf-8")
            for _, prow in prix_df.iterrows():
                bc = str(prow.get("BarCode", "")).strip()
                px = str(prow.get("Prix",    "")).strip().replace(",", ".")
                if bc and bc not in prix_lookup:
                    try:
                        prix_lookup[bc] = float(px)
                    except ValueError:
                        pass
            print(f"    Prix de gros : {len(prix_lookup):,} références indexées")
        except Exception as exc:
            print(f"⚠️  Prix de gros illisible ({exc}) — prix à 0.00")
    else:
        print("⚠️  'prix de gros.csv' absent — prix à 0.00")

    # ── Prix d'achat BDD (pour WEB / client 164) : ref → float ──────────
    bdd_prix_lookup: dict[str, float] = {}
    for ref, (desig_bdd, _) in bdd_ref_only.items():
        pass  # on va plutôt reconstruire depuis bdd_df
    # Reconstruction depuis les données BDD déjà lues
    if bdd_path.exists():
        try:
            _bdd_prix_df = pd.read_csv(
                bdd_path, sep="\t", dtype=str, low_memory=False, encoding="latin1"
            )
            _bdd_prix_df["_ref"] = (
                _bdd_prix_df["Reference_Article"].fillna("").str.strip().str.lstrip("% .")
            )
            for _, brow in _bdd_prix_df.iterrows():
                ref = brow["_ref"]
                px  = str(brow.get("Prix_achat", "")).strip().replace(",", ".")
                if ref and ref not in bdd_prix_lookup:
                    try:
                        bdd_prix_lookup[ref] = float(px)
                    except ValueError:
                        pass
            print(f"    Prix BDD (WEB)  : {len(bdd_prix_lookup):,} références indexées")
        except Exception as exc:
            print(f"⚠️  Prix BDD illisible ({exc}) — prix WEB à 0.00")

    # ── Construction des lignes ──────────────────────────────────────────
    HEADER = [
        "Numéro de commande", "Date", "CLIENT", "Commande",
        "escompte", "filler", "port", "frais postaux", "adlivraison",
        "devise", "date de liv sou", "date de liv prev",
        "Référence PRECO FW15", "filler", "BarCode", "COULEUR", "Taille",
        "DESIGNATION", "QUANTITE", "PU", "remise", "total HT",
        "TOTAL TAXES", "taux devises", "filler", "filler",
    ]

    rows: list[list[str]] = []   # chaque élément = liste de champs (sans numéro de ligne)

    boutiques_sans_num: list[str] = []
    boutiques_exported: set[str]  = set()

    for _, row in proposed.iterrows():
        boutique = str(row.get("boutique", "")).strip()
        barcode  = str(row.get("barcode",  "")).strip()
        taille   = str(row.get("taille",   "")).strip()
        try:
            qte = int(float(str(row.get("qte_proposee", 0))))
        except (ValueError, TypeError):
            qte = 0

        if qte <= 0:
            continue

        num = mag_to_num.get(boutique) or mag_to_num.get(str(boutique).strip().upper())
        if num is None:
            if boutique not in boutiques_sans_num:
                boutiques_sans_num.append(boutique)
            continue

        ordre_num = f"REASUCC{date_str}{num}"

        # Designation + Couleur
        # Priorité couleur : stock CENTRAL (couleur réellement en stock) > BDD
        desig, couleur = bdd_lookup.get(
            (barcode, taille),
            bdd_lookup.get((barcode, ""), bdd_ref_only.get(barcode, ("", "")))
        )
        # Écrase la couleur par celle du stock CENTRAL si disponible
        # (évite les articles multi-couleurs où la BDD prend la mauvaise)
        couleur_central = central_couleur.get((barcode, taille), "")
        if couleur_central:
            couleur = couleur_central
        # Sanitize : supprimer tabs/sauts de ligne qui corrompraient le TSV
        desig  = desig.replace("\t", " ").replace("\r", "").replace("\n", " ").strip()
        couleur = couleur.replace("\t", " ").replace("\r", "").replace("\n", " ").strip()

        # Prix : BDD (Prix_achat) pour WEB (164), prix de gros pour les autres
        if str(boutique).strip().upper() == "WEB":
            pu = bdd_prix_lookup.get(barcode, 0.0)
        else:
            pu = prix_lookup.get(barcode, 0.0)
        pu_str = f"{pu:.2f}"

        rows.append([
            ordre_num,   # Numéro de commande
            date_str,    # Date
            str(num),    # CLIENT
            "BL",        # Commande
            "0",         # escompte
            "",          # filler (col 5)
            "0",         # port
            "0",         # frais postaux
            "LIV" if str(boutique).strip().upper() == "WEB" else "LIVRAISON",  # adlivraison
            "EUR",       # devise
            deliv_date,  # date de liv sou
            deliv_date,  # date de liv prev
            ref_batch,   # Référence PRECO FW15
            "",          # filler (col 13)
            barcode,     # BarCode
            couleur,     # COULEUR
            taille,      # Taille
            desig,       # DESIGNATION
            str(qte),    # QUANTITE
            pu_str,      # PU
            "0",         # remise
            "0",         # total HT
            "0",         # TOTAL TAXES
            "1",         # taux devises
            "",          # filler (col 24)
            "",          # filler (col 25)
        ])
        boutiques_exported.add(boutique)


    # ── Écriture du fichier ──────────────────────────────────────────
    lines: list[str] = []
    lines.append("\t".join(HEADER))
    for fields in rows:
        lines.append("\t".join(fields))

    content = "\n".join(lines) + "\n"
    with open(output_path, "w", encoding="latin1", errors="replace") as fout:
        fout.write(content)

    return len(rows), len(boutiques_exported), boutiques_sans_num


# ──────────────────────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────────────────────


def main() -> None:
    args    = parse_args()
    use_gui = not args.no_gui and not (args.stock and args.ventes)

    central_src = None
    if use_gui:
        gui      = pick_files_gui(args.reserve)
        stock_p  = gui["stock"]
        ventes_p = gui["ventes"]
        output_p = gui["output"]
        reserve  = gui["reserve"]
        central_src = gui.get("central")
    else:
        if not args.stock or not args.ventes:
            raise ValueError("--stock et --ventes sont requis en mode --no-gui")
        stock_p  = Path(args.stock)
        ventes_p = Path(args.ventes)
        output_p = Path(args.output) if args.output else Path(f"reassort_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
        reserve  = args.reserve
        central_src = Path(args.stock_central) if args.stock_central else None

    # ── Mise à jour du stock CENTRAL de référence (si un export est fourni) ─
    if central_src:
        print(f"📦  Chargement de l'export stock CENTRAL : {Path(central_src).name}")
        update_stock_central(central_src)

    # Copie automatique BDD Fastmag
    # Source configurable via la variable d'environnement FASTMAG_SRC.
    # Sinon, on cherche le fichier dans le dossier du script (cas Mac/local).
    import os as _os
    _env_src = _os.environ.get("FASTMAG_SRC")
    if _env_src:
        FASTMAG_SRC = Path(_env_src)
    else:
        FASTMAG_SRC = Path(__file__).parent / "fastmag_bdd.xls"
    FASTMAG_DST = Path(__file__).parent / "fastmag_bdd.xls"
    FASTMAG_DST_TXT = Path(__file__).parent / "fastmag_bdd.xls.txt"
    if FASTMAG_SRC.exists():
        import shutil
        shutil.copy2(FASTMAG_SRC, FASTMAG_DST)
        print(f"📋  BDD Fastmag copiée : {FASTMAG_DST.name} ({FASTMAG_DST.stat().st_size // 1024} Ko)")
    elif FASTMAG_DST.exists() or FASTMAG_DST_TXT.exists():
        local = FASTMAG_DST if FASTMAG_DST.exists() else FASTMAG_DST_TXT
        print(f"📋  BDD Fastmag : version locale disponible ({local.name})")
    else:
        print(f"⚠️  BDD Fastmag introuvable : {FASTMAG_SRC}")

    SEP = "─" * 56
    print(f"\n{'═'*56}")
    print("  AGENT RÉASSORT MULTI-BOUTIQUES")
    print(f"{'═'*56}")
    print(f"  Stock      : {stock_p.name}")
    print(f"  Ventes     : {ventes_p.name}")
    print(f"  Réserve    : {reserve}")
    print(f"  Couverture : {args.target_days} jours")
    print(f"{'═'*56}\n")

    print("⏳  Chargement des données...")
    stock_df  = load_stock(stock_p)
    ventes_df, sales_start, sales_end, sales_days = load_ventes(ventes_p, args.sales_days)

    # ── Remplacement du stock CENTRAL par l'export Fastmag ────────────────
    # Stock CENTRAL : export Fastmag à placer dans le dossier du script
    _central_dst = Path(__file__).parent / "stock_central.xls"
    if _central_dst.exists():
        try:
            _df_p = pd.read_csv(_central_dst, sep='\t', encoding='latin1', dtype=str)
            _df_p.columns = [c.strip() for c in _df_p.columns]
            _ref_col = next(c for c in _df_p.columns if c.lower() in ('référence', 'reference', 'réf'))
            _df_p["barcode"] = _df_p[_ref_col].str.strip()
            _df_p["taille"]  = _df_p["Taille"].str.strip()
            _df_p["stock"]   = pd.to_numeric(
                _df_p["Stock"].str.replace(',', '.', regex=False), errors='coerce'
            ).fillna(0)
            _df_p["boutique"] = reserve
            _central_rows = _df_p[["boutique", "barcode", "taille", "stock"]].copy()
            # Remplacer les lignes CENTRAL dans stock_df
            stock_df = pd.concat([
                stock_df[stock_df["boutique"] != reserve],
                _central_rows,
            ], ignore_index=True)
            print(f"    Stock CENTRAL Fastmag : {len(_central_rows):,} lignes chargées ✓")
        except Exception as _e:
            print(f"    ⚠️  stock_central.xls illisible ({_e}) — stock CENTRAL du fichier principal conservé")

    print(f"    Stock  : {len(stock_df):>7,} lignes  |  {stock_df['boutique'].nunique()} boutiques (dont {reserve})")
    print(f"    Ventes : {len(ventes_df):>7,} lignes  |  période {sales_start} → {sales_end}  ({sales_days}j)")

    if reserve not in stock_df["boutique"].values:
        available = ", ".join(sorted(stock_df["boutique"].unique()))
        raise ValueError(f"Boutique réserve '{reserve}' introuvable.\nBoutiques disponibles : {available}")

    reserve_total = int(stock_df[stock_df["boutique"] == reserve]["stock"].sum())
    print(f"    Réserve {reserve} : {reserve_total:,} pièces disponibles\n")

    print("🤖  L'agent analyse les besoins...")
    promo_df = compute_promo_rates(ventes_df, threshold=args.promo_seuil)
    n_promo  = int(promo_df["promo_alerte"].sum()) if not promo_df.empty else 0
    if n_promo:
        print(f"    Filtre promo : {n_promo} articles exclus (remise > {int(args.promo_seuil*100)}%)")

    promo_exemptions = (
        set(b.strip().upper() for b in args.promo_exemptions.split(","))
        if args.promo_exemptions else PROMO_EXEMPTIONS
    )
    taille_exemptions = (
        set(b.strip().upper() for b in args.taille_exemptions.split(","))
        if args.taille_exemptions else {"WEB"}
    )

    proposed, risk, top, promo_exclus, taille_seule_exclus = run_agent(
        stock_df, ventes_df, reserve, args.target_days, sales_days,
        promo_df=promo_df, promo_seuil=args.promo_seuil,
        promo_exemptions=promo_exemptions,
        min_taille_share=args.min_taille_share,
        taille_exemptions=taille_exemptions,
        min_qty_core=args.min_qty_core,
        min_covered_mens=args.min_covered_mens,
        min_covered_shoes=args.min_covered_shoes,
        min_saison_year=args.min_saison_year,
        saison_exemptions=None,
        min_stock_floor=args.min_stock_floor,
        floor_min_sales=args.floor_min_sales,
    )

    run_date_now  = datetime.now()
    script_dir    = Path(__file__).parent

    # ── Exclusions manuelles de références (avant Excel ET Fastmag) ──────
    refs_set: set[str] = set()
    refs_exclues_txt  = script_dir / "refs_exclues.txt"
    refs_exclues_xlsx = script_dir / "refs_exclues.xlsx"
    if refs_exclues_txt.exists():
        try:
            refs_set = set(
                l.strip().upper()
                for l in refs_exclues_txt.read_text(encoding="utf-8").splitlines()
                if l.strip()
            )
        except Exception as e:
            print(f"  ⚠️  Impossible de lire refs_exclues.txt : {e}")
    elif refs_exclues_xlsx.exists():
        try:
            df_excl = pd.read_excel(refs_exclues_xlsx, dtype=str)
            col_ref = next(
                (c for c in df_excl.columns if c.strip().lower() in ("reference", "référence", "ref")),
                None
            )
            if col_ref:
                refs_set = set(df_excl[col_ref].dropna().str.strip().str.upper().tolist())
        except Exception as e:
            print(f"  ⚠️  Impossible de lire refs_exclues.xlsx : {e}")
    if refs_set and not proposed.empty:
        mask_excl = proposed["barcode"].str.strip().str.upper().isin(refs_set)
        nb_refs_exclues = int(mask_excl.sum())
        proposed = proposed[~mask_excl].copy()
        print(f"  📋  Refs exclues chargées : {len(refs_set)} — {nb_refs_exclues} lignes filtrées")

    total_qte = int(proposed["qte_proposee"].sum()) if not proposed.empty else 0

    print(f"\n{SEP}")
    print(f"  📦  Transferts proposés    : {len(proposed):,} lignes")
    print(f"  📊  Quantité à transférer  : {total_qte:,} pièces")
    print(f"  ⚠️   Risques sans réserve   : {len(risk):,} références")
    if not promo_exclus.empty:
        print(f"  🚫  Exclus promo           : {len(promo_exclus):,} besoins ignorés")
    if not taille_seule_exclus.empty:
        print(f"  👗  Tailles isolées        : {len(taille_seule_exclus):,} transferts exclus")

    if not proposed.empty:
        print("\n  Détail par priorité :")
        for prio, grp in proposed.groupby("priorite"):
            print(f"    {prio:30s}  {int(grp['qte_proposee'].sum()):>5,} pcs  ({len(grp)} refs)")
        print("\n  Top 5 boutiques :")
        top_b = proposed.groupby("boutique")["qte_proposee"].sum().sort_values(ascending=False).head(5)
        for b, q in top_b.items():
            print(f"    {b:20s}  {int(q):>5,} pcs")

    print(f"{SEP}\n")

    # Export Excel
    print("💾  Génération du fichier Excel...")
    build_excel(
        output_p, proposed, risk, top, promo_exclus, taille_seule_exclus,
        reserve, sales_start, sales_end, args.target_days, sales_days,
    )
    boutiques_n = stock_df[stock_df["boutique"] != reserve]["boutique"].nunique()
    print(f"\n✅  Fichier généré : {output_p}")
    print(f"    Feuilles : Synthèse | Tous transferts | {boutiques_n} boutiques | Alertes | Tailles | Risques | Top ventes\n")

    # ── PDF par magasin (réassort à recevoir) + zip pour l'email ──────────
    print("📄  Génération des PDF par magasin...")
    pdf_zip = None
    try:
        # S'assure que reportlab est présent (installe tout seul au besoin).
        try:
            import reportlab  # noqa: F401
        except ImportError:
            import subprocess
            print("    ⏳  reportlab manquant — installation automatique...")
            _ok = False
            for _flags in (["--user", "--break-system-packages"], ["--user"],
                           ["--break-system-packages"], []):
                try:
                    subprocess.check_call(
                        [sys.executable, "-m", "pip", "install", "-q", *_flags, "reportlab"]
                    )
                    import importlib
                    importlib.invalidate_caches()
                    import reportlab  # noqa: F401
                    _ok = True
                    break
                except Exception:
                    continue
            if not _ok:
                raise ImportError("reportlab")

        from generer_pdf_magasins import generer_pdfs
        pdf_dir = Path(output_p).parent / f"PDF_{Path(output_p).stem}"
        pdfs = generer_pdfs(output_p, stock_central=script_dir / "stock_central.xls", out_dir=pdf_dir)
        print(f"    {len(pdfs)} PDF générés dans : {pdf_dir}")
        # Zip de tous les PDF → pièce jointe de l'email récap
        if pdfs:
            import zipfile
            pdf_zip = Path(output_p).parent / f"PDF_{Path(output_p).stem}.zip"
            with zipfile.ZipFile(pdf_zip, "w", zipfile.ZIP_DEFLATED) as zf:
                for p in pdfs:
                    zf.write(p, arcname=Path(p).name)
            print(f"    📦  Zip pour l'email : {pdf_zip.name}\n")
    except ImportError:
        print("    ⚠️  PDF non générés — reportlab n'a pas pu être installé.")
        print(f"       Lance à la main :  {sys.executable} -m pip install --user --break-system-packages reportlab\n")
    except Exception as _e_pdf:
        import traceback
        print(f"    ⚠️  PDF non générés — erreur : {_e_pdf}")
        traceback.print_exc()
        print()

    # Génération Fastmag
    fastmag_out   = script_dir / f"IMPORT_FASTMAG_FOOTKORNER_{run_date_now.strftime('%d%m%Y')}.txt"
    nb_lignes     = 0
    if not proposed.empty:
        print("📤  Génération fichier import Fastmag...")
        nb_lignes, nb_boutiques, sans_num = generate_fastmag_import(
            proposed, fastmag_out, script_dir, run_date=run_date_now
        )
        if nb_lignes > 0:
            sep2 = "─" * 56
            print(f"\n{sep2}")
            print(f"  📋  CONTRÔLE IMPORT FASTMAG")
            print(f"{sep2}")
            print(f"  Fichier         : {fastmag_out.name}")
            print(f"  Lignes export   : {nb_lignes:,}")
            print(f"  Boutiques       : {nb_boutiques}")
            listing_ctrl = pd.read_excel(script_dir / "Footkorner - Listing mag.xlsx", dtype=str)
            mag_num_map  = dict(zip(
                listing_ctrl["code mag fastmag"].fillna("").str.strip(),
                listing_ctrl["NUM"].fillna("").str.strip()
            ))
            boutique_totals = (
                proposed[proposed["qte_proposee"].astype(float) > 0]
                .groupby("boutique")["qte_proposee"]
                .apply(lambda x: int(x.astype(float).sum()))
                .sort_values(ascending=False)
            )
            total_fm = sum(qty for b, qty in boutique_totals.items() if mag_num_map.get(str(b), "?") != "?")
            print(f"  Pièces totales  : {total_fm:,}")
            print(f"\n  Détail par boutique :")
            for bout, qty in boutique_totals.items():
                num_str = mag_num_map.get(str(bout), "?")
                if num_str != "?":
                    print(f"    [{num_str:>3}] {str(bout):20s}  {qty:>5,} pcs")
            if sans_num:
                print(f"\n  ⚠️  Boutiques sans numéro Fastmag : {', '.join(sans_num)}")
            print(f"{sep2}\n")
            print(f"✅  Import Fastmag  : {fastmag_out}")
        else:
            print("⚠️  Aucune ligne exportée (vérifier listing mag et BDD Fastmag)")

    # Email récap
    if not proposed.empty:
        print("\n📧  Envoi du mail récap...")
        html_body = build_email_html(
            proposed, risk, promo_exclus, taille_seule_exclus,
            sales_start, sales_end, args.target_days, sales_days, reserve,
            run_date=run_date_now,
        )
        fastmag_path = fastmag_out if nb_lignes > 0 else None
        send_email_recap(html_body, output_p, fastmag_path, script_dir,
                         run_date=run_date_now, pdf_zip_path=pdf_zip)

    if use_gui:
        try:
            import tkinter as tk
            from tkinter import messagebox
            root = tk.Tk(); root.withdraw()
            messagebox.showinfo(
                "Agent Reassort",
                f"Analyse terminée !\n\nTransferts : {len(proposed):,} lignes\n"
                f"Quantité   : {total_qte:,} pièces\nRisques    : {len(risk):,}\n\n"
                f"Fichier : {output_p}"
            )
            root.destroy()
        except Exception:
            pass


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        import traceback, datetime
        log_path = Path.cwd() / f"erreur_reassort_{datetime.date.today():%Y%m%d}.txt"
        with open(log_path, "w", encoding="utf-8") as lf:
            traceback.print_exc(file=lf)
        print(f"\n❌  Erreur fatale — détails dans {log_path}")
        traceback.print_exc()
    finally:
        input("\nAppuie sur Entrée pour fermer...")
