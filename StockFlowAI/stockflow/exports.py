"""Module exports - Generation de l'export Excel (7 onglets).

1. Transferts recommandes
2. Synthese par flux
3. Simulation avant/apres
4. Propositions d'implantation
5. Cas non traites
6. Parametres
7. Journal d'execution
"""

from __future__ import annotations

import math
from pathlib import Path
from typing import Dict, List

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

_LETTER_ORDER = {"XXS": 0, "XS": 1, "S": 2, "M": 3, "L": 4, "XL": 5,
                 "XXL": 6, "XXXL": 7, "3XL": 7, "TU": 9}


def _size_sort_key(t: str):
    t = str(t).strip().upper()
    if t in _LETTER_ORDER:
        return (0, _LETTER_ORDER[t], t)
    try:
        return (1, float(t.replace(",", ".")), t)
    except ValueError:
        return (2, 0.0, t)


def _fmt_dispo(sizes: Dict[str, float]) -> str:
    """Formate {taille: qte} trie logiquement : 'S:3 · M:5 · L:2'."""
    items = sorted(sizes.items(), key=lambda kv: _size_sort_key(kv[0]))
    return " · ".join(f"{t}:{int(round(q))}" for t, q in items) if items else "-"


def enrich_dispo(transfers: pd.DataFrame, base: pd.DataFrame, stock_final: dict) -> pd.DataFrame:
    """Ajoute aux transferts la disponibilite par taille chez le destinataire,
    avant (stock initial) et apres (stock final, tous transferts appliques),
    pour la reference concernee. Donne une vue complete de la courbe de tailles.
    """
    if transfers is None or transfers.empty:
        return transfers
    avant: Dict[tuple, Dict[str, float]] = {}
    for r in base.itertuples(index=False):
        q = float(getattr(r, "stock_actuel", 0) or 0)
        if q > 0:
            avant.setdefault((str(r.magasin), str(r.reference)), {})[str(r.taille)] = q
    apres: Dict[tuple, Dict[str, float]] = {}
    for (mag, ref, coul, taille), q in stock_final.items():
        if q > 0:
            apres.setdefault((str(mag), str(ref)), {})[str(taille)] = q
    t = transfers.copy()
    t["dispo_avant_dest"] = [_fmt_dispo(avant.get((str(r.destinataire), str(r.reference)), {}))
                             for r in t.itertuples(index=False)]
    t["dispo_finale_dest"] = [_fmt_dispo(apres.get((str(r.destinataire), str(r.reference)), {}))
                              for r in t.itertuples(index=False)]
    return t
from openpyxl.utils import get_column_letter

from .parameters import Parameters


# libelles finaux (colonnes) pour l'onglet transferts
TRANSFER_LABELS = {
    "priorite": "Priorite",
    "score": "Score",
    "expediteur": "Expediteur",
    "destinataire": "Destinataire",
    "reference": "Reference (code-barre)",
    "couleur": "Couleur",
    "taille": "Taille",
    "quantite": "Quantite",
    "stock_exp_avant": "Stock expediteur avant",
    "stock_exp_apres": "Stock expediteur apres",
    "cov_exp_avant": "Couv. expediteur avant",
    "cov_exp_apres": "Couv. expediteur apres",
    "stock_dest_avant": "Stock destinataire avant",
    "stock_dest_apres": "Stock destinataire apres",
    "cov_dest_avant": "Couv. destinataire avant",
    "cov_dest_apres": "Couv. destinataire apres",
    "grille_avant": "Grille avant",
    "grille_apres": "Grille apres",
    "dispo_avant_dest": "Dispo destinataire avant (par taille)",
    "dispo_finale_dest": "Dispo destinataire finale (par taille)",
    "picking_prevu": "Reassort Picking prevu",
    "besoin_residuel": "Besoin residuel",
    "motif": "Motif du transfert",
    "distance_km": "Distance estimee (km)",
    "destination_num": "Destination n0",
}


def build_flux_summary(transfers: pd.DataFrame, base: pd.DataFrame,
                       params: Parameters) -> pd.DataFrame:
    """Onglet 2 - synthese par flux (expediteur -> destinataire)."""
    cols = ["expediteur", "destinataire", "nb_references", "nb_tailles", "nb_pieces",
            "valeur_stock", "score_moyen", "priorite", "nb_colis_estime"]
    if transfers is None or transfers.empty:
        return pd.DataFrame(columns=cols)

    prix = base.drop_duplicates(["reference", "couleur", "taille"]).set_index(
        ["reference", "couleur", "taille"])["prix_vente"] if "prix_vente" in base else None

    def _valeur(group: pd.DataFrame) -> float:
        total = 0.0
        for t in group.itertuples(index=False):
            pv = 0.0
            if prix is not None:
                try:
                    pv = float(prix.loc[(t.reference, t.couleur, t.taille)])
                except Exception:
                    pv = 0.0
            total += t.quantite * pv
        return total

    pieces_par_colis = 30  # estimation logistique simple
    rows = []
    for (exp, dest), g in transfers.groupby(["expediteur", "destinataire"]):
        pieces = float(g["quantite"].sum())
        score_moy = round(float(g["score"].mean()), 1)
        rows.append({
            "expediteur": exp,
            "destinataire": dest,
            "nb_references": int(g["reference"].nunique()),
            "nb_tailles": int(len(g)),
            "nb_pieces": pieces,
            "valeur_stock": round(_valeur(g), 0),
            "score_moyen": score_moy,
            "priorite": _priorite_from_score(score_moy, params),
            "nb_colis_estime": max(1, math.ceil(pieces / pieces_par_colis)),
        })
    return pd.DataFrame(rows).sort_values("score_moyen", ascending=False).reset_index(drop=True)


def _priorite_from_score(score: float, params: Parameters) -> str:
    from .parameters import classer_score
    return classer_score(score)


def build_cas_non_traites(optimizer_result, needs: pd.DataFrame,
                          quality_report, base: pd.DataFrame,
                          web_codes) -> pd.DataFrame:
    """Onglet 5 - cas non traites (blocages, besoins sans donneur, anomalies)."""
    rows: List[Dict] = []

    # blocages du moteur
    for b in optimizer_result.blocked:
        rows.append({
            "categorie": "Transfert bloque",
            "magasin": b.get("magasin", ""),
            "reference": b.get("reference", ""),
            "couleur": b.get("couleur", ""),
            "taille": b.get("taille", ""),
            "detail": b.get("motif_blocage", ""),
        })

    # besoins non couverts (qte_restante > 0)
    if needs is not None and not needs.empty and "qte_restante" in needs:
        reste = needs[needs["qte_restante"] > 0]
        for r in reste.itertuples(index=False):
            rows.append({
                "categorie": "Besoin non couvert",
                "magasin": str(r.magasin),
                "reference": str(r.reference),
                "couleur": str(r.couleur),
                "taille": str(r.taille),
                "detail": f"Besoin residuel {getattr(r, 'qte_restante', 0):.0f} ({getattr(r, 'type_besoin', '')})",
            })

    # references sans vente (dormant / stock sans ventes)
    if base is not None and not base.empty:
        sans_vente = base[(base["moyenne_quotidienne"] <= 0) & (base["stock_actuel"] > 0)]
        agg = sans_vente.groupby(["magasin", "reference"], as_index=False)["stock_actuel"].sum()
        for r in agg.head(500).itertuples(index=False):
            rows.append({
                "categorie": "Reference sans vente (dormant)",
                "magasin": str(r.magasin),
                "reference": str(r.reference),
                "couleur": "",
                "taille": "",
                "detail": f"{r.stock_actuel:.0f} pieces sans vente sur la periode",
            })
        # magasins sans ville
        if "ville" in base:
            sans_ville = base[base["ville"].astype(str).str.strip().isin(["", "nan", "None"])]
            for mag in sorted(sans_ville["magasin"].astype(str).unique()):
                rows.append({"categorie": "Magasin sans ville", "magasin": mag,
                             "reference": "", "couleur": "", "taille": "",
                             "detail": "Proximite geo non exploitable"})

    # anomalies qualite non bloquantes
    if quality_report is not None:
        for a in quality_report.anomalies:
            rows.append({
                "categorie": f"Anomalie donnee ({a.severite})",
                "magasin": "",
                "reference": "",
                "couleur": "",
                "taille": "",
                "detail": f"[{a.fichier}] {a.message}",
            })

    cols = ["categorie", "magasin", "reference", "couleur", "taille", "detail"]
    if not rows:
        return pd.DataFrame(columns=cols)
    return pd.DataFrame(rows)[cols]


def build_journal(journal: Dict) -> pd.DataFrame:
    rows = [{"cle": k, "valeur": v} for k, v in journal.items()]
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Ecriture Excel avec mise en forme
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1a1d29")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
PRIO_FILLS = {
    "Prioritaire": PatternFill("solid", fgColor="C6EFCE"),
    "Fortement recommande": PatternFill("solid", fgColor="D9EAD3"),
    "Recommande": PatternFill("solid", fgColor="FFF2CC"),
    "A valider": PatternFill("solid", fgColor="FCE5CD"),
}


def _write_sheet(writer, sheet_name: str, df: pd.DataFrame, prio_col: str | None = None):
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.sheets[sheet_name]
    for col_idx, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # largeur auto approx
        maxlen = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str).head(200)])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(45, max(10, maxlen + 2))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    # coloration priorite
    if prio_col and prio_col in df.columns:
        pcol = list(df.columns).index(prio_col) + 1
        for r in range(len(df)):
            val = df.iloc[r][prio_col]
            fill = PRIO_FILLS.get(str(val))
            if fill:
                ws.cell(row=r + 2, column=pcol).fill = fill


def write_fiche_revue(path: str | Path, transfers: pd.DataFrame,
                      marque_map: dict | None = None, top: int | None = None) -> Path:
    """Genere une 'fiche de revue' epuree pour validation par les equipes.

    Tableau simple trie par score, avec deux colonnes vides a remplir :
    « OK ? » (menu deroulant OK/NON/?) et « Commentaire ».
    """
    from openpyxl.worksheet.datavalidation import DataValidation

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    t = transfers.copy() if transfers is not None else pd.DataFrame()

    if not t.empty:
        t = t.sort_values("score", ascending=False).reset_index(drop=True)
        if top:
            t = t.head(top)
        if marque_map:
            key = list(zip(t["reference"].astype(str), t["couleur"].astype(str)))
            t["marque"] = [marque_map.get(k, "") for k in key]
        fiche = pd.DataFrame({
            "N0": range(1, len(t) + 1),
            "Priorite": t.get("priorite"),
            "Score": t.get("score"),
            "Marque": t.get("marque", ""),
            "Expediteur": t.get("expediteur"),
            "Destinataire": t.get("destinataire"),
            "Reference (code-barre)": t.get("reference"),
            "Taille": t.get("taille"),
            "Quantite": t.get("quantite"),
            "Couv. dest. avant": t.get("cov_dest_avant"),
            "Couv. dest. apres": t.get("cov_dest_apres"),
            "Grille avant": t.get("grille_avant"),
            "Grille apres": t.get("grille_apres"),
            "Dispo dest. avant (par taille)": t.get("dispo_avant_dest"),
            "Dispo dest. finale (par taille)": t.get("dispo_finale_dest"),
            "Reassort Picking prevu": t.get("picking_prevu"),
            "Motif": t.get("motif"),
            "OK ?": "",
            "Commentaire": "",
        })
    else:
        fiche = pd.DataFrame({"Info": ["Aucun transfert a revoir"]})

    # mode d'emploi
    guide = pd.DataFrame({
        "Fiche de revue - mode d'emploi": [
            "Objectif : valider les transferts recommandes par StockFlow AI.",
            "1. Parcourez les lignes (deja triees du plus prioritaire au moins prioritaire).",
            "2. Colonne « OK ? » : choisissez OK / NON / ? (menu deroulant).",
            "3. Colonne « Commentaire » : precisez la raison d'un refus ou un ajustement.",
            "4. La colonne « Motif » explique pourquoi chaque transfert est propose.",
            "",
            "Critere de validation (brief) : au moins 80% des 50 premieres lignes",
            "doivent etre jugees pertinentes par le metier.",
            "",
            "Rappel : aucun transfert n'est execute automatiquement - vous gardez la main.",
        ]
    })

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        _write_sheet(writer, "Revue", fiche, prio_col="Priorite")
        guide.to_excel(writer, sheet_name="Mode d'emploi", index=False)
        # habillage colonnes a remplir + menu deroulant
        if not fiche.empty and "OK ?" in fiche.columns:
            ws = writer.sheets["Revue"]
            cols = list(fiche.columns)
            n = len(fiche)
            fill = PatternFill("solid", fgColor="FFF2CC")
            for name in ("OK ?", "Commentaire"):
                ci = cols.index(name) + 1
                letter = get_column_letter(ci)
                ws.column_dimensions[letter].width = 16 if name == "OK ?" else 40
                for r in range(2, n + 2):
                    ws.cell(row=r, column=ci).fill = fill
            ci = cols.index("OK ?") + 1
            letter = get_column_letter(ci)
            dv = DataValidation(type="list", formula1='"OK,NON,?"', allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{letter}2:{letter}{n + 1}")
        # mise en forme feuille guide
        gws = writer.sheets["Mode d'emploi"]
        gws.column_dimensions["A"].width = 80
        gws.cell(row=1, column=1).font = Font(bold=True, size=12)
    return path


def export_excel(path: str | Path, *, transfers: pd.DataFrame, flux: pd.DataFrame,
                 simulation_global: pd.DataFrame, simulation_stores: pd.DataFrame,
                 implantations: pd.DataFrame, cas_non_traites: pd.DataFrame,
                 parametres: pd.DataFrame, journal: pd.DataFrame,
                 top_references: pd.DataFrame | None = None,
                 criticite: pd.DataFrame | None = None) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    # onglet transferts : renommage lisible
    t = transfers.copy() if transfers is not None else pd.DataFrame()
    if not t.empty:
        # couleur embarquee dans le code-barre => colonne vide : on la retire
        if "couleur" in t.columns and t["couleur"].astype(str).str.strip().replace(
                {"nan": "", "None": ""}).eq("").all():
            t = t.drop(columns=["couleur"])
        ordered = [c for c in TRANSFER_LABELS if c in t.columns]
        extra = [c for c in t.columns if c not in TRANSFER_LABELS]
        t = t[ordered + extra].rename(columns=TRANSFER_LABELS)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        _write_sheet(writer, "1-Transferts", t if not t.empty else pd.DataFrame({"Info": ["Aucun transfert retenu"]}),
                     prio_col="Priorite")
        _write_sheet(writer, "2-Synthese flux", flux if not flux.empty else pd.DataFrame({"Info": ["Aucun flux"]}),
                     prio_col="priorite")
        _write_sheet(writer, "3-Simulation", simulation_global)
        if simulation_stores is not None and not simulation_stores.empty:
            _write_sheet(writer, "3b-Simulation magasins", simulation_stores)
        _write_sheet(writer, "4-Implantations",
                     implantations if not implantations.empty else pd.DataFrame({"Info": ["Aucune proposition"]}))
        _write_sheet(writer, "5-Cas non traites",
                     cas_non_traites if not cas_non_traites.empty else pd.DataFrame({"Info": ["Aucun cas"]}))
        _write_sheet(writer, "6-Parametres", parametres)
        _write_sheet(writer, "7-Journal", journal)
        if top_references is not None and not top_references.empty:
            _write_sheet(writer, "8-Top references", top_references)
        if criticite is not None and not criticite.empty:
            _write_sheet(writer, "9-Criticite magasins", criticite)

    return path
